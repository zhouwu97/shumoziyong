"""独立复算 2024-C 问级 Result；不调用求解器目标函数。"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def read_excel_assignments(path: Path, data) -> list[dict[str, object]]:
    # 模板较小；普通模式支持高效随机读取，避免 read_only 的逐单元格重复扫描。
    workbook = load_workbook(path, read_only=False, data_only=True)
    result = []
    for year in range(2024, 2031):
        sheet = workbook[str(year)]
        occurrences = defaultdict(list)
        for row in range(2, sheet.max_row + 1):
            plot = sheet.cell(row, 2).value
            if plot in data["plots"]:
                occurrences[str(plot)].append(row)
        for row in range(2, sheet.max_row + 1):
            plot_id = sheet.cell(row, 2).value
            if plot_id not in data["plots"]:
                continue
            plot_type = data["plots"][plot_id]["type"]
            for column, crop_id in enumerate(range(1, 42), 3):
                value = sheet.cell(row, column).value
                if not isinstance(value, (int, float)) or abs(float(value)) <= 1e-8:
                    continue
                if crop_id == 16:
                    season = "单季"
                elif len(occurrences[str(plot_id)]) > 1 and row == occurrences[str(plot_id)][1]:
                    season = "第二季"
                elif plot_type in {"平旱地", "梯田", "山坡地"}:
                    season = "单季"
                else:
                    season = "第一季"
                result.append({"year": year, "plot_id": str(plot_id), "season": season, "crop_id": crop_id, "area_mu": float(value)})
    return result


def assignment_difference(left, right) -> float:
    def grouped(items):
        values = defaultdict(float)
        for item in items:
            values[(int(item["year"]), str(item["plot_id"]), str(item["season"]), int(item["crop_id"]))] += float(item["area_mu"])
        return values
    a, b = grouped(left), grouped(right)
    return max((abs(a[key] - b[key]) for key in set(a) | set(b)), default=0.0)


def passed(summary: str) -> dict[str, str]:
    return {"status": "passed", "summary": summary}


def failed(summary: str) -> dict[str, str]:
    return {"status": "failed", "summary": summary}


def independent_sample_profits(data, assignments, factors) -> np.ndarray:
    """独立实现随机情景利润，不调用求解模块的评价函数。"""
    count = factors["demand"].shape[0]
    production = defaultdict(lambda: np.zeros(count))
    costs = np.zeros((count, 7))
    for item in assignments:
        year = int(item["year"])
        year_index = year - 2024
        plot_id = str(item["plot_id"])
        season = str(item["season"])
        crop_id = int(item["crop_id"])
        area = float(item["area_mu"])
        plot_type = data["plots"][plot_id]["type"]
        stat = data["stats"][(plot_type, season, crop_id)]
        production[(year_index, season, crop_id)] += area * stat["yield"] * factors["yield"][:, year_index, crop_id - 1]
        costs[:, year_index] += area * stat["cost"] * factors["cost"][:, year_index, crop_id - 1]
    profits = -costs.sum(axis=1)
    for (year_index, season, crop_id), amount in production.items():
        demand = data["sales_2023"].get((crop_id, season), 0.0) * factors["demand"][:, year_index, crop_id - 1]
        price = data["price_by_crop_season"][(crop_id, season)] * factors["price"][:, year_index, crop_id - 1]
        profits += np.minimum(amount, demand) * price
    return profits


def risk_stats(values: np.ndarray) -> dict[str, float]:
    p05 = float(np.quantile(values, 0.05))
    return {"mean": float(values.mean()), "std": float(values.std(ddof=1)), "p05": p05, "cvar05": float(values[values <= p05].mean())}


def average_pair_correlation(values: np.ndarray) -> float:
    correlations = [float(np.corrcoef(values[:, :, left].ravel(), values[:, :, left + 1].ravel())[0, 1]) for left in range(0, 40, 2)]
    return float(np.mean(correlations))


def pair_correlation(values: np.ndarray, pairs) -> float:
    return float(np.mean([
        np.corrcoef(values[:, :, int(left) - 1].ravel(), values[:, :, int(right) - 1].ravel())[0, 1]
        for left, right in pairs
    ]))


def cvar(values: np.ndarray, probability: float) -> float:
    cutoff = float(np.quantile(values, probability))
    return float(values[values <= cutoff].mean())


def check_q1(run_dir: Path, question_dir: Path, validator) -> dict[str, object]:
    data = validator.load_problem_data(run_dir / "materials/attachments/附件1.xlsx", run_dir / "materials/attachments/附件2.xlsx")
    records = {}
    for label in ("q1_waste", "q1_discount"):
        records[label] = json.loads((question_dir / f"results/artifacts/{label}_plan.json").read_text(encoding="utf-8"))
    objective_diffs = []
    violations = []
    max_raw = 0.0
    for label, record in records.items():
        recomputed = validator.evaluate_objective(record["assignments"], data, label)
        objective_diffs.append(abs(recomputed - float(record["objective_reported"])))
        current, raw = validator.check_constraints(record["assignments"], data, 1e-5)
        violations.extend(current); max_raw = max(max_raw, raw)
    excel_diffs = [
        assignment_difference(records["q1_waste"]["assignments"], read_excel_assignments(run_dir / "official/result1_1.xlsx", data)),
        assignment_difference(records["q1_discount"]["assignments"], read_excel_assignments(run_dir / "official/result1_2.xlsx", data)),
    ]
    body = (question_dir / "paper.typ").read_text(encoding="utf-8")
    body_keys = ("q1-waste-profit", "q1-discount-profit", "q1-max-constraint-violation", "q1-max-mip-gap")
    gaps = [float(records[label]["solver"]["mip_gap"]) for label in records]
    return {
        "checks": {
            "hard_constraints": passed(f"独立复算约束违规数为 0，最大原始超限 {max_raw:.3g}") if not violations else failed(f"独立复算发现 {len(violations)} 项违规"),
            "objective_recalculation": passed(f"最大目标复算误差 {max(objective_diffs):.6g} 元") if max(objective_diffs) <= 1e-4 else failed(f"目标复算误差 {max(objective_diffs):.6g} 元"),
            "excel_readback": passed(f"官方 Excel 最大面积反读误差 {max(excel_diffs):.3g} 亩") if max(excel_diffs) <= 1e-5 else failed(f"Excel 面积误差 {max(excel_diffs):.3g} 亩"),
            "sales_scenarios": passed("浪费与半价两种销售情形均独立复算") if set(records) == {"q1_waste", "q1_discount"} else failed("销售情形不完整"),
            "mip_gap_disclosure": passed(f"正文披露最大 MIP gap={max(gaps):.6g}") if "MIP gap" in body else failed("正文未披露 MIP gap"),
            "body_metric_binding": passed("Q1 正文引用全部生成变量") if all(key in body for key in body_keys) else failed("Q1 正文缺少生成变量引用"),
        }
    }


def check_q2(run_dir: Path, question_dir: Path, validator) -> dict[str, object]:
    data = validator.load_problem_data(run_dir / "materials/attachments/附件1.xlsx", run_dir / "materials/attachments/附件2.xlsx")
    result = json.loads((question_dir / "results/result.json").read_text(encoding="utf-8"))
    selected = json.loads((question_dir / "results/artifacts/selected_plan.json").read_text(encoding="utf-8"))
    baseline = json.loads((run_dir / "questions/q1/results/artifacts/q1_waste_plan.json").read_text(encoding="utf-8"))
    violations, max_raw = validator.check_constraints(selected["assignments"], data, 1e-5)
    excel_diff = assignment_difference(selected["assignments"], read_excel_assignments(run_dir / "official/result2.xlsx", data))
    with np.load(question_dir / "results/final_evaluation_factors.npz") as archive:
        factors = {key: archive[key] for key in archive.files}
    selected_values = independent_sample_profits(data, selected["assignments"], factors)
    baseline_values = independent_sample_profits(data, baseline["assignments"], factors)
    selected_stats = risk_stats(selected_values)
    baseline_stats = risk_stats(baseline_values)
    metrics = result["metrics"]
    expected = {
        "final_mean_profit": selected_stats["mean"],
        "final_std_profit": selected_stats["std"],
        "final_p05_profit": selected_stats["p05"],
        "final_cvar05_profit": selected_stats["cvar05"],
        "baseline_mean_profit": baseline_stats["mean"],
        "relative_mean_improvement": (selected_stats["mean"] - baseline_stats["mean"]) / abs(baseline_stats["mean"]),
    }
    max_metric_diff = max(abs(float(metrics[key]["value"]) - value) for key, value in expected.items())
    final_count = int(metrics["final_sample_count"]["value"])
    split = json.loads((question_dir / "results/artifacts/sample_split.json").read_text(encoding="utf-8"))
    seeds = [*split["training"]["seeds"], split["candidate_evaluation"]["seed"], split["final_evaluation"]["seed"], split["bootstrap_seed"]]
    differences = pd.read_csv(question_dir / "results/tables/plan_difference.csv")
    solver = pd.read_csv(question_dir / "results/tables/solver_evidence.csv")
    uncertainty = pd.read_csv(question_dir / "results/tables/uncertainty_model.csv")
    convergence = pd.read_csv(question_dir / "results/tables/sample_convergence.csv")
    intervals = pd.read_csv(question_dir / "results/tables/bootstrap_intervals.csv")
    body = (question_dir / "paper.typ").read_text(encoding="utf-8")
    body_keys = tuple(f"q2-{key.replace('_', '-')}" for key in ("final_mean_profit", "final_p05_profit", "final_cvar05_profit", "paired_mean_improvement", "relative_mean_improvement", "plan_l1_change", "max_mip_gap", "final_sample_count"))
    max_gap = float(solver["mip_gap"].max())
    distinct = assignment_difference(selected["assignments"], baseline["assignments"])
    return {
        "checks": {
            "hard_constraints": passed(f"独立复算约束违规数为 0，最大原始超限 {max_raw:.3g}") if not violations else failed(f"独立复算发现 {len(violations)} 项违规"),
            "excel_readback": passed(f"result2.xlsx 最大面积反读误差 {excel_diff:.3g} 亩") if excel_diff <= 1e-5 else failed(f"Excel 面积误差 {excel_diff:.3g} 亩"),
            "held_out_evaluation": passed(f"独立复算 {final_count} 个最终情景，最大指标误差 {max_metric_diff:.6g}") if factors["demand"].shape[0] == final_count and max_metric_diff <= 1e-4 else failed(f"最终评估不一致，误差 {max_metric_diff:.6g}"),
            "sample_separation": passed("训练、候选评价、最终评估与 bootstrap seed 互不重叠") if len(seeds) == len(set(seeds)) and split.get("sets_disjoint_by_seed") is True else failed("样本阶段 seed 存在重叠"),
            "plan_reoptimisation": passed(f"Q2 与 Q1 最大单元面积差 {distinct:.3f} 亩，共 {len(differences)} 个变化单元") if distinct > 1e-5 and len(differences) > 0 else failed("Q2 未形成不同于 Q1 的规划"),
            "solver_evidence": passed(f"3 个候选均记录 incumbent/bound/终止原因，最大 gap={max_gap:.3%}") if len(solver) == 3 and max_gap <= 0.10 and {"solver_objective", "mip_dual_bound", "mip_gap", "message"}.issubset(solver.columns) else failed("Q2 求解证据不完整或 gap 超门槛"),
            "uncertainty_audit": passed("随机变量、分布区间和时间规则表完整") if len(uncertainty) >= 6 and {"variable", "distribution", "time_rule"}.issubset(uncertainty.columns) else failed("Q2 随机模型披露不完整"),
            "stability_analysis": passed("样本量收敛、风险权重与 bootstrap 区间均存在") if set(convergence["sample_count"].astype(int)) == {256, 512, 1024, 2048} and set(intervals["metric"]) == {"mean", "p05", "cvar05"} else failed("Q2 稳定性证据不完整"),
            "body_metric_binding": passed("Q2 正文引用全部生成变量") if all(key in body for key in body_keys) else failed("Q2 正文缺少生成变量引用"),
        }
    }


def check_q3(run_dir: Path, question_dir: Path, validator) -> dict[str, object]:
    data = validator.load_problem_data(run_dir / "materials/attachments/附件1.xlsx", run_dir / "materials/attachments/附件2.xlsx")
    result = json.loads((question_dir / "results/result.json").read_text(encoding="utf-8"))
    selected = json.loads((question_dir / "results/artifacts/selected_plan.json").read_text(encoding="utf-8"))
    q2 = json.loads((run_dir / "questions/q2/results/artifacts/selected_plan.json").read_text(encoding="utf-8"))
    violations, max_raw = validator.check_constraints(selected["assignments"], data, 1e-5)
    with np.load(question_dir / "results/final_evaluation_factors.npz") as archive:
        stored = {key: archive[key] for key in archive.files}
    correlated = {key: stored[f"correlated_{key}"] for key in ("demand", "yield", "cost", "price")}
    independent = {key: stored[f"independent_{key}"] for key in ("demand", "yield", "cost", "price")}
    selected_corr = independent_sample_profits(data, selected["assignments"], correlated)
    q2_corr = independent_sample_profits(data, q2["assignments"], correlated)
    selected_ind = independent_sample_profits(data, selected["assignments"], independent)
    paired = selected_corr - q2_corr
    corr_stats = risk_stats(selected_corr)
    q2_corr_stats = risk_stats(q2_corr)
    metrics = result["metrics"]
    expected = {
        "correlated_mean_profit": corr_stats["mean"],
        "correlated_std_profit": corr_stats["std"],
        "correlated_p05_profit": corr_stats["p05"],
        "correlated_cvar05_profit": corr_stats["cvar05"],
        "q2_correlated_mean_profit": q2_corr_stats["mean"],
        "q2_correlated_cvar05_profit": q2_corr_stats["cvar05"],
        "cvar05_improvement": corr_stats["cvar05"] - q2_corr_stats["cvar05"],
        "risk_utility_improvement": (0.75 * selected_corr.mean() + 0.25 * cvar(selected_corr, 0.10)) - (0.75 * q2_corr.mean() + 0.25 * cvar(q2_corr, 0.10)),
        "paired_mean_improvement": float(paired.mean()),
        "paired_improvement_probability": float(np.mean(paired > 0)),
        "independent_mean_profit": float(selected_ind.mean()),
    }
    max_metric_diff = max(abs(float(metrics[key]["value"]) - value) for key, value in expected.items())
    mechanism = json.loads((question_dir / "results/artifacts/correlation_mechanism.json").read_text(encoding="utf-8"))
    demand_price = float(np.corrcoef(stored["correlated_latent_demand"].ravel(), stored["correlated_latent_price"].ravel())[0, 1])
    substitution = pair_correlation(stored["correlated_latent_demand"], mechanism["substitution_pairs"])
    complement = pair_correlation(stored["correlated_latent_demand"], mechanism["complement_pairs"])
    mechanism_error = max(
        abs(demand_price - float(metrics["demand_price_correlation"]["value"])),
        abs(substitution - float(metrics["substitution_pair_correlation"]["value"])),
        abs(complement - float(metrics["complement_pair_correlation"]["value"])),
    )
    split = json.loads((question_dir / "results/artifacts/sample_split.json").read_text(encoding="utf-8"))
    seeds = [*split["training"]["seeds"], split["candidate_evaluation"]["seed"], split["final_correlated"]["seed"], split["final_independent"]["seed"]]
    differences = pd.read_csv(question_dir / "results/tables/plan_difference.csv")
    solver = pd.read_csv(question_dir / "results/tables/solver_evidence.csv")
    relations = pd.read_csv(question_dir / "results/tables/crop_relations.csv")
    body = (question_dir / "paper.typ").read_text(encoding="utf-8")
    body_keys = tuple(f"q3-{key.replace('_', '-')}" for key in ("correlated_mean_profit", "correlated_cvar05_profit", "paired_mean_improvement", "paired_improvement_probability", "plan_l1_change", "demand_price_correlation", "substitution_pair_correlation", "complement_pair_correlation", "max_mip_gap", "final_sample_count"))
    final_count = int(metrics["final_sample_count"]["value"])
    distinct = assignment_difference(selected["assignments"], q2["assignments"])
    q3_utility = 0.75 * selected_corr.mean() + 0.25 * cvar(selected_corr, 0.10)
    q2_utility = 0.75 * q2_corr.mean() + 0.25 * cvar(q2_corr, 0.10)
    mechanism_valid = mechanism.get("demand_price_loading") == -0.35 and mechanism.get("cost_market_loading") == 0.50 and demand_price < 0 and substitution < 0 and complement > 0
    return {
        "checks": {
            "hard_constraints": passed(f"独立复算约束违规数为 0，最大原始超限 {max_raw:.3g}") if not violations else failed(f"独立复算发现 {len(violations)} 项违规"),
            "held_out_evaluation": passed(f"独立复算相关/独立各 {final_count} 情景，最大指标误差 {max_metric_diff:.6g}") if correlated["demand"].shape[0] == independent["demand"].shape[0] == final_count and max_metric_diff <= 1e-4 else failed(f"最终评估不一致，误差 {max_metric_diff:.6g}"),
            "sample_separation": passed("训练、候选评价、相关最终与独立对照 seed 互不重叠") if len(seeds) == len(set(seeds)) and split.get("sets_disjoint_by_seed") is True else failed("Q3 样本阶段 seed 存在重叠"),
            "correlation_mechanism": passed(f"需求-价格={demand_price:.3f}，替代={substitution:.3f}，互补={complement:.3f}") if mechanism_valid and mechanism_error <= 1e-10 and len(relations) == 10 else failed("相关机制参数、方向、映射或结果绑定不一致"),
            "plan_reoptimisation": passed(f"Q3 与 Q2 最大单元面积差 {distinct:.3f} 亩，共 {len(differences)} 个变化单元") if distinct > 1e-5 and len(differences) > 0 else failed("Q3 未形成不同于 Q2 的规划"),
            "solver_evidence": passed("固定离散格局的三次面积再分配均达到 0 gap") if len(solver) == 3 and float(solver["mip_gap"].max()) <= 1e-9 and (solver["fixed_active_size"] == solver["support_size"]).all() else failed("Q3 精确再分配证据不完整"),
            "scenario_comparison": passed(f"最终相关集风险效用提高 {q3_utility-q2_utility:.2f} 元") if q3_utility > q2_utility and len(paired) == final_count else failed("Q3 最终相关集未显示风险效用改进"),
            "body_metric_binding": passed("Q3 正文引用全部生成变量") if all(key in body for key in body_keys) else failed("Q3 正文缺少生成变量引用"),
        }
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--question-dir", type=Path, required=True)
    args = parser.parse_args()
    sys.path.insert(0, str(args.run_dir / "shared"))
    import independent_validator as validator
    qid = args.question_dir.name.lower()
    if qid == "q1":
        report = check_q1(args.run_dir, args.question_dir, validator)
    elif qid == "q2":
        report = check_q2(args.run_dir, args.question_dir, validator)
    elif qid == "q3":
        report = check_q3(args.run_dir, args.question_dir, validator)
    else:
        report = {"checks": {"checker_execution": failed(f"{qid} Checker 尚未实现")}}
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
