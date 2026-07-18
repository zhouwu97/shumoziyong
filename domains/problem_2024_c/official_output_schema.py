"""官方结果工作簿的单元格合同、导出与反向读取。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook

from .data_model import YEARS, ProblemData


FIRST_SEASON_ROWS = range(2, 56)
SECOND_SEASON_ROWS = range(56, 84)
CROP_COLUMNS = range(3, 44)


@dataclass(frozen=True)
class Assignment:
    plot_id: str
    year: int
    season: str
    crop_id: int
    area_mu: float


@dataclass(frozen=True)
class WorkbookContract:
    sheet_names: tuple[str, ...]
    crop_names: tuple[str, ...]
    first_season_plots: tuple[str, ...]
    second_season_plots: tuple[str, ...]
    merged_ranges: tuple[str, ...]


def inspect_template(path: Path) -> WorkbookContract:
    book = load_workbook(path, data_only=False, read_only=False)
    if tuple(book.sheetnames) != tuple(str(year) for year in YEARS):
        raise ValueError(f"官方模板年份工作表不完整：{book.sheetnames}")
    first = book[book.sheetnames[0]]
    crop_names = tuple(str(first.cell(1, column).value).strip() for column in CROP_COLUMNS)
    first_plots = tuple(str(first.cell(row, 2).value).strip() for row in FIRST_SEASON_ROWS)
    second_plots = tuple(str(first.cell(row, 2).value).strip() for row in SECOND_SEASON_ROWS)
    if len(set(crop_names)) != 41 or len(set(first_plots)) != 54 or len(set(second_plots)) != 28:
        raise ValueError("官方模板作物列或地块行存在重复/缺失")
    return WorkbookContract(
        sheet_names=tuple(book.sheetnames),
        crop_names=crop_names,
        first_season_plots=first_plots,
        second_season_plots=second_plots,
        merged_ranges=tuple(sorted(str(item) for item in first.merged_cells.ranges)),
    )


def export_official_workbook(
    template_path: Path,
    output_path: Path,
    data: ProblemData,
    assignments: Iterable[Assignment | Mapping[str, object]],
) -> None:
    """保留模板结构和样式，仅写入规定的面积单元格。"""
    contract = inspect_template(template_path)
    expected_crops = tuple(data.crops[crop_id].name for crop_id in sorted(data.crops))
    if tuple(name.replace(" ", "") for name in contract.crop_names) != expected_crops:
        raise ValueError("官方模板作物列与附件 1 不一致")

    book = load_workbook(template_path, data_only=False, read_only=False)
    crop_columns = {crop_id: column for crop_id, column in zip(sorted(data.crops), CROP_COLUMNS)}
    first_rows = {plot_id: row for plot_id, row in zip(contract.first_season_plots, FIRST_SEASON_ROWS)}
    second_rows = {plot_id: row for plot_id, row in zip(contract.second_season_plots, SECOND_SEASON_ROWS)}
    for sheet in book.worksheets:
        for row in (*FIRST_SEASON_ROWS, *SECOND_SEASON_ROWS):
            for column in CROP_COLUMNS:
                sheet.cell(row, column).value = 0

    for raw in assignments:
        item = raw if isinstance(raw, Assignment) else Assignment(
            plot_id=str(raw["plot_id"]),
            year=int(raw["year"]),
            season=str(raw["season"]),
            crop_id=int(raw["crop_id"]),
            area_mu=float(raw.get("area_mu", raw.get("area", 0.0))),
        )
        if item.area_mu <= 1e-9:
            continue
        rows = second_rows if item.season == "第二季" else first_rows
        if item.plot_id not in rows:
            raise ValueError(f"模板不存在地块季次：{item.plot_id}-{item.season}")
        cell = book[str(item.year)].cell(rows[item.plot_id], crop_columns[item.crop_id])
        cell.value = round(float(cell.value or 0.0) + item.area_mu, 6)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)


def import_official_workbook(path: Path, data: ProblemData) -> list[Assignment]:
    """从官方结果表恢复非零决策变量，供独立 Validator 使用。"""
    contract = inspect_template(path)
    book = load_workbook(path, data_only=True, read_only=False)
    crop_ids = dict(zip(CROP_COLUMNS, sorted(data.crops)))
    result: list[Assignment] = []
    for year in YEARS:
        sheet = book[str(year)]
        for rows, season in ((FIRST_SEASON_ROWS, "第一季"), (SECOND_SEASON_ROWS, "第二季")):
            for row in rows:
                plot_id = str(sheet.cell(row, 2).value).strip()
                plot_type = data.plots[plot_id].plot_type
                for column in CROP_COLUMNS:
                    value = sheet.cell(row, column).value
                    if isinstance(value, (int, float)) and float(value) > 1e-9:
                        crop_id = crop_ids[column]
                        is_single = season == "第一季" and (
                            plot_type in {"平旱地", "梯田", "山坡地"}
                            or (plot_type == "水浇地" and crop_id == 16)
                        )
                        actual_season = "单季" if is_single else season
                        result.append(Assignment(plot_id, year, actual_season, crop_id, float(value)))
    return result
