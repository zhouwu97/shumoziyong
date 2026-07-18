"""读取并校验 2024-C 全部官方数据附件。"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data_model import Crop, CropStat, Planting2023, Plot, ProblemData, SourceCell


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]


def default_audit_output_path(repository_root: Path | None = None) -> Path:
    """在包含非 ASCII 字符的仓库路径下仍返回仓库内部的固定位置。"""
    root = (repository_root or REPOSITORY_ROOT).resolve()
    return root / "capability_evidence" / "2024_c_full_closure" / "a0_official_materials_audit.json"


def resolve_material_root() -> Path:
    """返回官方材料根目录；受控 CI 可通过环境变量替换。"""
    configured = os.environ.get("SHUMO_OFFICIAL_MATERIALS_DIR", "").strip()
    return Path(configured).expanduser().resolve() if configured else REPOSITORY_ROOT / "official_materials"


def _clean(value: object) -> str:
    return str(value).replace(" ", "").strip()


def _merged_values(sheet: Worksheet) -> dict[tuple[int, int], object]:
    result: dict[tuple[int, int], object] = {}
    for merged in sheet.merged_cells.ranges:
        value = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                result[(row, column)] = value
    return result


def _rows(sheet: Worksheet, start: int, end: int, columns: int) -> Iterator[tuple[int, list[object]]]:
    merged = _merged_values(sheet)
    for row in range(start, end + 1):
        yield row, [merged.get((row, column), sheet.cell(row, column).value) for column in range(1, columns + 1)]


def _source(workbook: Path, sheet: Worksheet, row: int, columns: int) -> SourceCell:
    return SourceCell(
        workbook=workbook.name,
        sheet=sheet.title,
        row=row,
        columns=tuple(f"{get_column_letter(column)}{row}" for column in range(1, columns + 1)),
    )


def _price_range(value: object) -> tuple[float, float]:
    parts = str(value).strip().split("-")
    if len(parts) != 2:
        raise ValueError(f"销售价格区间格式无效：{value!r}")
    return float(parts[0]), float(parts[1])


def load_problem_data(material_root: Path | None = None) -> ProblemData:
    """从附件 1、附件 2 构造规范对象，并拒绝静默缺行。"""
    root = (material_root or resolve_material_root()) / "2024_C" / "attachments"
    attachment_1 = root / "附件1.xlsx"
    attachment_2 = root / "附件2.xlsx"
    missing = [str(path) for path in (attachment_1, attachment_2) if not path.is_file()]
    if missing:
        raise FileNotFoundError("缺少 2024-C 官方附件：" + ", ".join(missing))

    book_1 = load_workbook(attachment_1, data_only=True, read_only=False)
    land_sheet = book_1["乡村的现有耕地"]
    crop_sheet = book_1["乡村种植的农作物"]

    plots: dict[str, Plot] = {}
    for row, values in _rows(land_sheet, 2, 55, 4):
        plot_id = _clean(values[0])
        plots[plot_id] = Plot(
            plot_id=plot_id,
            plot_type=_clean(values[1]),
            area_mu=float(values[2]),
            source=_source(attachment_1, land_sheet, row, 4),
        )

    crops: dict[int, Crop] = {}
    for row, values in _rows(crop_sheet, 2, 42, 5):
        crop_id = int(values[0])
        crop_type = _clean(values[2])
        crops[crop_id] = Crop(
            crop_id=crop_id,
            name=_clean(values[1]),
            crop_type=crop_type,
            is_legume="豆类" in crop_type,
            source=_source(attachment_1, crop_sheet, row, 5),
        )

    book_2 = load_workbook(attachment_2, data_only=True, read_only=False)
    planting_sheet = book_2["2023年的农作物种植情况"]
    stat_sheet = book_2["2023年统计的相关数据"]

    planting: list[Planting2023] = []
    for row, values in _rows(planting_sheet, 2, 88, 6):
        planting.append(
            Planting2023(
                plot_id=_clean(values[0]),
                crop_id=int(values[1]),
                season=_clean(values[5]),
                area_mu=float(values[4]),
                source=_source(attachment_2, planting_sheet, row, 6),
            )
        )

    statistics: dict[tuple[int, str, str], CropStat] = {}
    for row, values in _rows(stat_sheet, 2, 108, 8):
        low, high = _price_range(values[7])
        stat = CropStat(
            crop_id=int(values[1]),
            plot_type=_clean(values[3]),
            season=_clean(values[4]),
            yield_jin_per_mu=float(values[5]),
            cost_yuan_per_mu=float(values[6]),
            price_low_yuan_per_jin=low,
            price_high_yuan_per_jin=high,
            source=_source(attachment_2, stat_sheet, row, 8),
        )
        key = (stat.crop_id, stat.plot_type, stat.season)
        if key in statistics:
            raise ValueError(f"官方统计参数重复：{key}")
        statistics[key] = stat

    provisional = ProblemData(plots, crops, tuple(planting), statistics, {})
    expected_sales: dict[tuple[int, str], float] = {}
    for record in planting:
        plot_type = plots[record.plot_id].plot_type
        stat = provisional.stat(record.crop_id, plot_type, record.season)
        key = (record.crop_id, record.season)
        expected_sales[key] = expected_sales.get(key, 0.0) + record.area_mu * stat.yield_jin_per_mu

    if (len(plots), len(crops), len(planting), len(statistics)) != (54, 41, 87, 107):
        raise ValueError(
            "官方附件行数不完整："
            f"plots={len(plots)}, crops={len(crops)}, planting={len(planting)}, stats={len(statistics)}"
        )
    if abs(sum(item.area_mu for item in plots.values()) - 1213.0) > 1e-9:
        raise ValueError("地块总面积与附件 1 不一致")
    return ProblemData(plots, crops, tuple(planting), statistics, expected_sales)
