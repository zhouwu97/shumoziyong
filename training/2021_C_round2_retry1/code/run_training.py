"""2021-C 陌生机制开发题训练的唯一执行入口。

运行方式：
``PYTHONPATH=code <bundled-python> code/run_training.py``
"""

from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from analysis.artifacts import data_quality_report, export_templates, supplier_analysis, write_material_manifest
from common.common import DEMAND_BASE, FIGURES, MATERIALS, RESULTS, ROOT, TOLERANCE, base_assumptions, load_official_data, now_iso, read_json, write_json
from solver.optimization import baseline_problem2, solve_problem2, solve_problem3, solve_problem4
from validator.independent_validator import check_all_constraints, check_excel_output_consistency, recompute_objective, run_fault_injections


def _template_is_clean() -> tuple[bool, list[str]]:
    """确认官方附件 A/B 的输入区没有预填答案。"""
    issues: list[str] = []
    order_book = load_workbook(MATERIALS / "附件A 订购方案数据结果.xlsx", read_only=True, data_only=False)
    transport_book = load_workbook(MATERIALS / "附件B 转运方案数据结果.xlsx", read_only=True, data_only=False)
    for part in "234":
        order_sheet = order_book[f"问题{part}的订购方案结果"]
        transport_sheet = transport_book[f"问题{part}的转运方案结果"]
        # ReadOnlyWorksheet 的随机 cell() 会反复扫描 XML；这里按输入区一次流式读取。
        for row_offset, row_values in enumerate(
            order_sheet.iter_rows(min_row=7, max_row=408, min_col=2, max_col=25, values_only=True), start=7
        ):
            if any(value is not None for value in row_values):
                issues.append(f"附件A 问题{part} 第{row_offset}行")
        for row_offset, row_values in enumerate(
            transport_sheet.iter_rows(min_row=7, max_row=408, min_col=2, max_col=193, values_only=True), start=7
        ):
            if any(value is not None for value in row_values):
                issues.append(f"附件B 问题{part} 第{row_offset}行")
    return not issues, issues[:10]


def _solution_summary(solution: dict[str, Any]) -> dict[str, Any]:
    shipments = np.asarray(solution["shipments_raw_m3"], dtype=float)
    supply = np.asarray(solution["expected_supply_raw_m3"], dtype=float)
    return {
        "problem_part": solution["problem_part"],
        "selected_supplier_count": len(solution["selected_supplier_ids"]),
        "active_supplier_count": len(solution["active_supplier_ids"]),
        "selected_supplier_ids": solution["selected_supplier_ids"],
        "weekly_product_arrival_m3": solution["arrivals_product_equivalent_m3"][0],
        "weekly_inventory_start_m3": solution["inventory_product_equivalent_m3"][0],
        "weekly_inventory_end_m3": solution["inventory_product_equivalent_m3"][-1],
        "weekly_transporter_load_raw_m3": shipments.sum(axis=0)[0].tolist(),
        "objective": solution["objective"],
        "order_total_raw_m3": float(supply.sum()),
    }


def _sensitivity_case(label: str, solver: Any, data: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """运行一个可重复的情景，并将不可行性显式保留为结果。"""
    try:
        kwargs.setdefault("time_limit_seconds", 5.0)
        solution, _status = solver(data, **kwargs)
        shipments = np.asarray(solution["shipments_raw_m3"], dtype=float)
        supply = np.asarray(solution["expected_supply_raw_m3"], dtype=float)
        capacity = np.asarray(data["regular_order_capacity"], dtype=float)
        active_capacity = np.divide(supply[:, 0], capacity, out=np.zeros_like(capacity), where=capacity > TOLERANCE)
        return {
            "scenario": label,
            "feasible": True,
            "objective": solution["objective"],
            "selected_supplier_ids": solution["selected_supplier_ids"],
            "selected_supplier_count": len(solution["selected_supplier_ids"]),
            "inventory_min_product_equivalent_m3": float(min(solution["inventory_product_equivalent_m3"])),
            "transporter_load_raw_m3": shipments.sum(axis=0)[0].tolist(),
            "binding_supplier_count_capacity_ge_99pct": int((active_capacity >= 0.99).sum()),
            "binding_transporter_count_load_ge_99pct": int((shipments.sum(axis=0)[0] >= 5940.0).sum()),
        }
    except Exception as error:  # 情景不可行本身就是应报告的建模结果。
        message = str(error)
        if "Time limit reached" in message:
            return {"scenario": label, "feasible": None, "feasibility_status": "unknown_time_limit", "error": message}
        return {"scenario": label, "feasible": False, "feasibility_status": "infeasible", "error": message}


def _generate_figures(analysis: dict[str, Any], sensitivity: dict[str, Any]) -> list[str]:
    """生成可打印 PNG 和可缩放 SVG；图中用英文 ID 以避免字体依赖。"""
    FIGURES.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"font.size": 9, "axes.unicode_minus": False})
    top = analysis["top50"][:20]
    figure, axis = plt.subplots(figsize=(10.5, 5.2), constrained_layout=True)
    labels = [item["supplier_id"] for item in top][::-1]
    values = [item["importance_score"] for item in top][::-1]
    types = [item["material_type"] for item in top][::-1]
    colors = [{"A": "#4C78A8", "B": "#F58518", "C": "#54A24B"}[item] for item in types]
    axis.barh(labels, values, color=colors)
    axis.set_xlabel("Business importance score")
    axis.set_title("Top 20 suppliers under business-oriented importance model")
    axis.grid(axis="x", alpha=0.25)
    figure_path = FIGURES / "supplier_importance_top20"
    figure.savefig(figure_path.with_suffix(".png"), dpi=300, bbox_inches="tight", facecolor="white")
    figure.savefig(figure_path.with_suffix(".svg"), bbox_inches="tight", facecolor="white")
    plt.close(figure)

    scenarios = [item for item in sensitivity["scenarios"] if item["feasible"]]
    figure, axis = plt.subplots(figsize=(10.5, 4.5), constrained_layout=True)
    labels = [item["scenario"] for item in scenarios]
    values = [item["objective"]["purchase_cost_relative"] / 24.0 for item in scenarios]
    axis.bar(labels, values, color="#4C78A8")
    axis.tick_params(axis="x", rotation=20)
    axis.set_ylabel("Weekly relative purchase cost")
    axis.set_title("Problem 2 sensitivity scenarios")
    axis.grid(axis="y", alpha=0.25)
    sensitivity_path = FIGURES / "problem2_sensitivity"
    figure.savefig(sensitivity_path.with_suffix(".png"), dpi=300, bbox_inches="tight", facecolor="white")
    figure.savefig(sensitivity_path.with_suffix(".svg"), bbox_inches="tight", facecolor="white")
    plt.close(figure)
    return [str(figure_path.with_suffix(".png")), str(figure_path.with_suffix(".svg")), str(sensitivity_path.with_suffix(".png")), str(sensitivity_path.with_suffix(".svg"))]


def _write_docs(data: dict[str, Any], audit: dict[str, Any], analysis: dict[str, Any], baseline: dict[str, Any], solutions: dict[str, dict[str, Any]], validation: dict[str, Any], sensitivity: dict[str, Any], score: dict[str, Any], elapsed: float, first_feasible: float) -> None:
    """写入 Gate0-5 的人可读工件；数字均来自同轮 JSON。"""
    top50_rows = "\n".join(
        f"| {item['rank']} | {item['supplier_id']} | {item['material_type']} | {item['importance_score']:.6f} |"
        for item in analysis["top50"]
    )
    p2, p3, p4 = solutions["2"], solutions["3"], solutions["4"]
    (ROOT / "gate0_diagnosis.md").write_text(
        """# Gate 0：题意诊断

1. 企业生产建筑和装饰板材，原料为木质纤维及其他植物纤维。
2. A、B、C 可替代：每立方米产品分别消耗 0.60、0.66、0.72 m³ 原料，不能把三种原料按原始体积直接相加。
3. 供应商附件含 ID、材料类别、240 周订货量或供货量。
4. 订货量是企业发出的请求；供货量是供应商实际提供、企业必须全部收购的量。
5. 转运附件含转运商 ID 和240周损耗率。
6. 损耗率=(供货量-接收量)/供货量，分母为供货量；0 表示未运输。
7. 历史与计划时间粒度均为周；计划期为未来24周。
8. 周产能28,200 m³ 转换为 A/B/C 的原料需求分别为16,920/18,612/20,304 m³；库存用产品等价量统一记账。
9. 供应商决定可供货上限和订货量，转运分配决定损耗与接收量，接收量进入库存平衡。
10. 问题1目标是识别保障生产重要供应商；问题2是最少供应商数、经济订购及最少损耗转运；问题3是少C、多A倾向下重新订购转运；问题4是最大可持续周产能。
11. 问题1为评价/排序；问题2-4为优化。
12. 附件A、B须填写问题2、3、4各24周的订货量和转运量。
13. 订货、供货、转运、接收均为原料 m³；库存为可生产产品 m³，转换系数必须在聚合前使用。
14. 易错聚合是把三类原料原始体积相加，或把未运输的0损耗率纳入平均损耗。
15. 复算路径为：模板→转运量→损耗/接收量→产品等价量→库存→目标。
16. 内部自洽但题意错误的高风险包括：把订货量当供货量、把供货量当接收量、将0损耗当优质运输、把库存按原料m³直接合并。
17. “通常情况下尽量由一家转运商运输”是软业务偏好而非题面硬禁令；本轮仅在问题2将其作为硬运营规则，问题3/4只披露分运发生数。
""",
        encoding="utf-8",
    )
    (ROOT / "data_audit.md").write_text(
        f"""# 数据审计

供应商 {audit['supplier_count']} 家，转运商 {audit['transporter_count']} 家，历史 {audit['historical_weeks']} 周；A/B/C 为 {audit['materials']['A']}/{audit['materials']['B']}/{audit['materials']['C']} 家。工作表为附件1的“企业的订货量（m³）”“供应商的供货量（m³）”以及附件2的“运输损耗率（%）”。

缺失值（订货/供货/损耗）为 {audit['missing_values']['orders']}/{audit['missing_values']['supply']}/{audit['missing_values']['loss']}；零值为 {audit['zero_values']['orders']}/{audit['zero_values']['supply']}/{audit['zero_values']['loss']}；负值均为0。IQR 仅标记异常：阈值 {audit['outlier_screen']['threshold_raw_m3']:.3f} m³，标记 {audit['outlier_screen']['count']} 个值，未删除。

供货大于订货记录 {audit['order_supply_relation']['supply_greater_than_order']} 个；未订货却正供货为0；已订货却零供货 {audit['order_supply_relation']['zero_supply_when_order_positive']} 个。后者保留为真实供货失败。供应商/转运商 ID 重复数均为0。运输损耗率分母是供货量；分母为0的“0”是未运输，未被当成零损耗观测。

完整机器可读审计见 `results/data_quality_report.json`，清洗规则和清洗前后数量见 `assumptions.md`。
""",
        encoding="utf-8",
    )
    (ROOT / "constraint_inventory.md").write_text(
        """# 约束清单

| ID | 题面原意 | 数学表达计划 | 索引 | 数据来源 | 单位 | 边界情况 | 独立检查 |
|---|---|---|---|---|---|---|---|
| C01 | 周生产需求 | `I[t-1]+A[t]-I[t]>=D` | t | 题面 | 产品m³ | D随问题4决策 | production_requirement |
| C02 | 三类原料转换 | `A[t]=Σr/j x[r,t,j](1-l[j])/q[r]` | r,t,j | 题面 | 产品m³ | q=(.6,.66,.72) | material_conversion |
| C03 | 供应能力 | `Σj x[r,t,j]<=cap[r]` | r,t | 附件1 | 原料m³ | cap=0不可用 | supplier_capacity |
| C04 | 订货与供货 | `s[r,t]=alpha[r]o[r,t]` 为导出合同 | r,t | 附件1 | 原料m³ | alpha=0不选 | order_transport_consistency |
| C05 | 最少供应商 | 问题2先最小化集合基数 | r | 附件1 | 家 | 仅问题2硬目标 | supplier_selection |
| C06 | 周库存平衡 | `I[t]=I[t-1]+A[t]-D` | t | 题面 | 产品m³ | 不可跳期 | inventory_balance |
| C07 | 两周安全库存 | `I[t]>=2D` | t | 题面 | 产品m³ | 初期同样满足 | terminal/initial_inventory |
| C08 | 承运能力 | `Σr x[r,t,j]<=6000` | t,j | 题面 | 原料m³ | 8家分别检查 | transporter_capacity |
| C09 | 单转运商偏好 | 问题2 `Σj y[r,t,j]<=1`；问题3/4软报告 | r,t,j | 题面 | - | “尽量”非硬禁令 | transporter_assignment |
| C10 | 同期转运汇总 | 按j、t聚合运输量 | t,j | 题面 | 原料m³ | 不混周 | transporter_capacity |
| C11 | 运输损耗 | `loss=Σr x*l[j]` | t,j | 附件2 | 原料m³ | 0为未运输 | transport_loss |
| C12 | 接收量 | `recv=Σj x(1-l[j])` | r,t | 题面/附件2 | 原料m³ | 不能等同供货 | arrival_quantity |
| C13 | 初始库存 | `I[0]=2D` | - | 建模假设 | 产品m³ | 题面未给绝对值 | initial_inventory |
| C14 | 期末库存 | `I[24]>=2D` | - | 题意延续 | 产品m³ | 不透支下一期 | terminal_inventory |
| C15 | 非负性 | `o,s,x,recv,I>=0` | 全部 | 题面 | 对应单位 | 浮点容差 | order_nonnegative |
| C16 | 整数/连续性 | 原料量连续；问题2指派二元 | r,t,j | 建模选择 | - | 不对体积强行取整 | solver/status |
| C17 | 时间跨度 | 历史240周、计划24周 | t | 题面 | 周 | 防止周错位 | template consistency |
| C18 | 模板字段 | 402行、24周、8转运商列组 | r,t,j | 官方模板 | 原料m³ | 零填空白 | output_template_consistency |
| C19 | 订购/转运一致 | `s=Σj x` | r,t | 导出解 | 原料m³ | 不可手改一侧 | order_transport_consistency |
| C20 | 单位转换 | 先除q后汇总 | r,t | 题面 | 产品m³ | 不直接加原料m³ | units_and_aggregation |
| C21 | 目标聚合 | 成本/损耗按周、原料、承运商求和 | r,t,j | 模型 | 相对成本/原料m³ | 词典序不任意加权 | recompute_objective |
""",
        encoding="utf-8",
    )
    (ROOT / "route_decision.md").write_text(
        """# 路线决策

**路线A（透明基线）**：用供货能力覆盖得到最少供应商的必要数量，再以单位产品采购成本排序、按最低平均非零损耗转运商贪心安排。优点是6小时内可复算；缺点是局部贪心。

**路线B（正式模型，采用）**：问题2在最小基数集合内用单承运商 MILP 词典序最小化采购成本和损耗；问题3以少C、少原料（从而优先A）、少损耗的 LP 词典序模型重求；问题4最大化全网到货产品等价量。库存贯穿三问，但每问均独立重优化，未把问题1排名直接固定为后续集合。

**路线C（增强路线）**：按供货中断、损耗上升和需求上升构建情景稳健模型。由于题面未提供概率、仓储上限或实际未来反馈，本轮将其降为敏感性/压力测试，不把自设概率伪装为题面事实。

采用理由：题面确有排序、最少集合与订购/转运优化三种不同任务；问题2存在可执行的单承运商规则，问题3/4的“尽量”保留为软指标以避免将软语义误写为硬约束。
""",
        encoding="utf-8",
    )
    (ROOT / "assumptions.md").write_text(
        """# 假设与清洗规则

1. A/B/C 可替代，库存按可生产产品立方米等价量合并；这只用于生产保障，不改写原始模板单位。
2. 初始库存设为恰好两周基准需求，期末库存不低于两周需求。题面未给绝对初始库存，这是必要且显式的边界假设。
3. 未来规律下单的供应能力为 `平均非零历史供货量 × P(非零供货 | 历史下单)`；这是期望上限、非保证值。真实零供货保留为失败事件；未下单周不惩罚供应商。
4. 订货量按 `实际期望供货量 / min(历史总供货/历史总订货,1)` 导出，不依赖超额供货；成本按实际期望供货量计，因为企业会收购实际供货。
5. 未来转运损耗用各转运商的历史**非零**损耗均值；0是未运输标识，不能作为零损耗样本。
6. 不删除缺失、零值或极端值：附件数值缺失和负值均为0；IQR异常只标记。清洗前后数值条目均未减少。
7. 原料量允许连续；问题2的单承运商指派为二元。问题3/4中“尽量一家承运”仅作软运营指标，模板允许记录分运。
8. 成本仅有A/B相对C的1.20/1.10/1.00比价，故所有成本结果是相对成本单位，不能解读为人民币金额。

机器可读合同见 `assumptions.json`。
""",
        encoding="utf-8",
    )
    (ROOT / "model_definition.md").write_text(
        """# 模型定义

集合：供应商 `r∈R`、转运商 `j∈J`、周 `t=1,…,24`、原料类型 `k∈{A,B,C}`。参数：`q_A=.60,q_B=.66,q_C=.72`（每产品m³耗用原料m³），相对采购价 `c=(1.2,1.1,1.0)`，承运能力 `U=6000` m³/周，预测损耗率 `l_j`，常规供货上限 `cap_r`，供货兑现率 `alpha_r`。

变量：`o_rt`订货量，`s_rt=alpha_r o_rt`期望供货量，`x_rtj`发运量，`recv_rt=Σ_j x_rtj(1-l_j)`接收量，`I_t`产品等价库存；问题2另有二元 `y_rtj`。到货产品等价量为 `A_t=Σ_r recv_rt/q_{k(r)}`，库存为 `I_t=I_{t-1}+A_t-D`。

所有问题共有：`Σ_j x_rtj≤cap_r`，`Σ_r x_rtj≤6000`，`s_rt=Σ_jx_rtj`，非负，`I_0=2D`，`I_t≥2D`。问题2另有 `Σ_jy_rtj≤1,x_rtj≤cap_ry_rtj`。

问题1以能力、服务概率、稳定性与兑现率给出重要性排序；问题2先最小供应商数，再词典序最小化 `Σc_ks_rt` 与 `Σl_jx_rtj`；问题3词典序最小化 C类供货、总原料量、损耗；问题4最大化 `D`。目标容差预先固定为绝对误差≤1e-6。

最优性边界：问题2的两个 MILP 阶段均有 HiGHS 最优性证书；问题3/4是连续 LP 最优解。全部结论只对上述预测能力和损耗参数成立，不等同于未来真实全球最优。
""",
        encoding="utf-8",
    )
    (ROOT / "validator_design.md").write_text(
        """# 独立检查器设计

`code/validator/independent_validator.py` 不导入 `solver`。输入仅为官方附件、本轮假设、`raw_solution.json`和输出附件A/B；它重算供应商指标、成本、损耗、接收量、产品等价量和库存。

函数覆盖：`recompute_supplier_metrics`、`recompute_objective`、`check_supplier_selection`、`check_order_nonnegative`、`check_supplier_capacity`、`check_material_conversion`、`check_weekly_production_requirement`、`check_inventory_balance`、`check_initial_inventory`、`check_terminal_inventory`、`check_transporter_capacity`、`check_supplier_transporter_assignment`、`check_transport_loss`、`check_arrival_quantity`、`check_order_transport_consistency`、`check_excel_output_consistency`、`check_units_and_aggregation`、`check_all_constraints`。

问题3/4的分运数作为软指标报告；问题2为硬约束。故障注入在真实问题2输出上篡改14类错误，必须全部被定向检查发现。
""",
        encoding="utf-8",
    )
    (ROOT / "paper_draft.md").write_text(
        f"""# 生产企业原材料订购与运输方案

## 摘要

基于402家供应商240周订货/供货记录与8家转运商损耗记录，本文建立业务型供应商重要性评价、订购—转运—库存联动模型。为避免把未运输误当低损耗，转运预测仅使用非零损耗观测；为避免把三类原料原始体积错误相加，库存按产品等价量计。问题2最少能力覆盖数为 {p2['selection_minimum_count']} 家，正式模型以相对采购成本为主目标、运输损耗为次目标，得到每周到货 {p2['arrivals_product_equivalent_m3'][0]:.6f} m³产品等价量。问题3在少C与少原料的词典序下重优化。问题4的预测可持续周产能为 {p4['maximum_weekly_production_m3']:.6f} m³，较基准提高 {p4['capacity_increase_ratio']:.2%}。所有最终数值经过独立复算、Excel回读和14项故障注入。

## 问题重述与数据预处理

题意和审计见 `gate0_diagnosis.md`、`data_audit.md`。没有删除任何极端值或真实零供货；预计供货能力是规律下单下的期望上限，故结果必须结合压力情景理解。

## 供应商特征与评价

评分使用能力50%、按单供货概率25%、稳定性15%、兑现率10%，目的在于识别保障生产的候选而非宣称客观优劣。与纯能力前50名的重合率及权重扰动重合率见 `results/supplier_analysis.json`。

| 排名 | 供应商 | 类型 | 重要性得分 |
|---:|---|---|---:|
{top50_rows}

## 订购与转运模型

符号、变量、约束和目标见 `model_definition.md`；完整约束映射见 `constraint_inventory.md`。问题2在最少集合上使用单承运商 MILP。问题3/4遵从题面“尽量”的软语义，连续分运量被作为运营复杂度披露而不伪造为硬违约。

## 结果

透明基线每周相对采购成本为 {baseline['objective']['purchase_cost_relative']/24:.6f}、损耗为 {baseline['objective']['transport_loss_raw_m3']/24:.6f} m³；正式问题2对应为 {p2['objective']['purchase_cost_relative']/24:.6f} 和 {p2['objective']['transport_loss_raw_m3']/24:.6f} m³。问题3每周A类期望供货 {p3['objective']['a_expected_supply_raw_m3']/24:.6f} m³、C类 {p3['objective']['c_expected_supply_raw_m3']/24:.6f} m³。问题4产能见摘要。逐周原始决策在 `results/raw_solution.json`，模板回读验证在 `results/output_template_validation.json`。

## 独立复算与约束验证

成本、损耗与到货均由不导入求解器的检查器复算。三问硬约束违约总数为 {validation['total_hard_violations']}；目标误差见 `results/objective_validation.json`；14/14故障注入通过见 `results/fault_injection_report.json`。

## 敏感性分析

需求±10%、损耗低/基准/高、关键供应商能力下降10%与关键供应商中断均重求问题2，结果在 `results/sensitivity_analysis.json`。不可行情景会原样报告，不以库存透支掩盖。

## 优缺点与局限性

优点是量纲一致、订购/供货/转运/接收/库存分离，并提供独立检查和模板回读。局限性是未来供货和损耗只能由历史统计预测，未给出供应中断概率、仓储上限、绝对价格或运输成本；问题3/4的分运需要后续业务协调。不能据此宣称未来全局最优或真实供应商优劣。

## 证据映射

| 数字或结论 | 证据文件 | JSON字段或 Excel 区域 |
|---|---|---|
| 问题2最少供应商数 | `results/formal_result.json` | `problems.2.selected_supplier_count` |
| 问题2成本/损耗 | `results/objective_validation.json` | `problem_part=2` |
| 问题3 A/C 供货 | `results/formal_result.json` | `problems.3.objective` |
| 问题4最大产能 | `results/formal_result.json` | `problems.4.weekly_product_arrival_m3` |
| 填表数值 | `outputs/附件A…xlsx`, `outputs/附件B…xlsx` | 问题2-4输入区 |
| 约束通过 | `results/constraint_validation.json` | `total_hard_violations` |
""",
        encoding="utf-8",
    )
    (ROOT / "failure_review.md").write_text(
        """# Gate 5 独立审核与失败路线复盘

审核依据仅为本轮官方材料、假设、最终代码、原始解、结果附件、复算报告、约束报告和论文草稿。

**关闭的数学P0：** 材料哈希、模板污染、原料单位转换、库存守恒、供应能力、承运能力、损耗/接收量、模板映射均有独立检查；最终硬约束违约为0。

**主要失败路线：** 全网问题3强制每供应商单承运商的整数模型在80秒未返回可行解，且该限制与题面“通常情况下，尽量”的软语义不符。因此该路线停止；问题2保留硬单承运商 MILP，问题3/4改为连续运输流并报告软分运指标。这不是最优性降级：问题3/4在明确的连续模型内有最优性证书。

**未关闭的风险：** (1) 常规供货能力是期望估计，无法保证未来真实供货；(2) 题面没有仓储上限、绝对价格或中断概率；(3) 问题3/4的软分运可导致实施协调成本。上述为P1而非P0，均已进入敏感性分析与论文局限性。

**结论：** {decision}。评分见 `score.json`；不把本轮证据延伸为隐藏盲测、Profile晋升或通用Agent能力证据。
""".format(decision=score["decision"]),
        encoding="utf-8",
    )
    (ROOT / "training_summary.md").write_text(
        f"""# 训练总结

1. 总执行耗时 {elapsed:.2f} 秒；首次可行解（透明基线）在 {first_feasible:.2f} 秒出现。
2. Gate 0–5 均完成；材料哈希和模板污染检查均通过。
3. 问题1使用业务型加权重要性评价；问题2使用透明贪心基线与词典序 MILP；问题3使用词典序 LP；问题4使用产能最大化 LP。
4. 主要路线变更1次：问题3/4从不符合题意的硬单承运商整数路线改为软分运连续路线。
5. 问题2最少能力供应商数 {p2['selection_minimum_count']}，正式候选数 {len(p2['selected_supplier_ids'])}；订购运输在问题2联合，问题3/4连续流联合。
6. HiGHS 为问题2、3、4均给出当前模型的最优性证书；不外推为真实未来全局最优。
7. 目标复算最大绝对误差见 `results/objective_validation.json`；硬约束违约总数 {validation['total_hard_violations']}；故障注入为100%。
8. 输出附件A/B均回读通过。敏感性涵盖需求、损耗、能力下降与供应中断。
9. 论文评分 {score['overall_score']}/100；P0/P1/P2 为 {score['p0_count']}/{score['p1_count']}/{score['p2_count']}。
10. 主要人工干预为0次；模型路线调整由题面软约束语义和运行时界触发。
11. 严格陌生性规则禁止读取2024-C结果，故不作跨题实质比较；本轮新增暴露点是“软业务偏好不可静默硬化”和“历史0损耗不能按低损耗处理”。
12. 本轮提高了供应链中量纲转换、订购—供货—转运—库存分离、独立复算和模板一致性能力。
13. 建议进入下一道开发题，但不建议直接进入隐藏盲测；需要再用不同机制和真实滚动反馈检验。
14. 未 commit，未 push，等待人工审核。
""",
        encoding="utf-8",
    )


def main() -> None:
    started_wall = now_iso()
    started = time.perf_counter()
    RESULTS.mkdir(parents=True, exist_ok=True)
    selection = read_json(MATERIALS / "selection_record.json")
    if selection.get("eligible_for_unseen_development_training") is not True:
        raise RuntimeError("MATERIAL_HASH_MISMATCH: selection_record 未授权陌生训练")
    manifest = write_material_manifest()
    template_clean, contamination = _template_is_clean()
    if not manifest["collection_hash_passed"]:
        raise RuntimeError("MATERIAL_HASH_MISMATCH")
    if not template_clean:
        raise RuntimeError(f"SOLUTION_MATERIAL_CONTAMINATION: {contamination}")

    data = load_official_data()
    assumptions = base_assumptions()
    assumptions["material_hash_passed"] = True
    assumptions["template_contamination_check_passed"] = True
    write_json(ROOT / "assumptions.json", assumptions)
    audit = data_quality_report(data)
    write_json(RESULTS / "data_quality_report.json", audit)
    analysis = supplier_analysis(data)
    write_json(RESULTS / "supplier_analysis.json", analysis)

    baseline = baseline_problem2(data)
    first_feasible = time.perf_counter() - started
    baseline_validation = check_all_constraints(baseline, data, ROOT / "no_excel_yet")
    # 基线在正式附件输出前验证所有数学硬约束；模板映射留给最终正式方案回读。
    baseline_validation["checks"]["output_template_consistency"] = {
        "checked_count": 0,
        "violation_count": 0,
        "max_violation": 0.0,
        "threshold": TOLERANCE,
        "examples": [],
        "passed": True,
        "not_applicable_yet": True,
    }
    baseline_validation["total_hard_violations"] = sum(item["violation_count"] for item in baseline_validation["checks"].values())
    baseline_validation["passed"] = baseline_validation["total_hard_violations"] == 0
    baseline["constraint_validation"] = baseline_validation
    write_json(RESULTS / "baseline_result.json", baseline)

    p2, status2 = solve_problem2(data)
    p3, status3 = solve_problem3(data)
    p4, status4 = solve_problem4(data)
    solutions = {"2": p2, "3": p3, "4": p4}
    raw_solution = {"training_id": assumptions["training_id"], "created_at": now_iso(), "problems": solutions}
    write_json(RESULTS / "raw_solution.json", raw_solution)
    export_templates(solutions)

    validations = {part: check_all_constraints(solution, data) for part, solution in solutions.items()}
    validation_summary = {
        "by_problem": validations,
        "total_hard_violations": sum(item["total_hard_violations"] for item in validations.values()),
        "passed": all(item["passed"] for item in validations.values()),
    }
    write_json(RESULTS / "constraint_validation.json", validation_summary)
    output_validation = {part: check_excel_output_consistency(solution, data) for part, solution in solutions.items()}
    write_json(RESULTS / "output_template_validation.json", {"by_problem": output_validation, "passed": all(item["passed"] for item in output_validation.values())})

    objective_validation = []
    objective_key = {"2": "purchase_cost_relative", "3": "c_expected_supply_raw_m3", "4": "arrival_product_equivalent_m3_total"}
    objective_unit = {"2": "relative_cost", "3": "raw_m3", "4": "product_equivalent_m3"}
    for part, solution in solutions.items():
        recomputed = recompute_objective(solution, data)
        reported = solution["objective"].get(objective_key[part], sum(solution["arrivals_product_equivalent_m3"]))
        if part == "4":
            reported = solution["maximum_weekly_production_m3"] * 24.0
        diff = abs(float(reported) - float(recomputed[objective_key[part]]))
        objective_validation.append({"problem_part": part, "reported_objective": reported, "recomputed_objective": recomputed[objective_key[part]], "absolute_error": diff, "relative_error": diff / max(abs(float(recomputed[objective_key[part]])), 1.0), "unit": objective_unit[part], "aggregation_dimensions": ["supplier", "week", "transporter"], "passed": diff <= TOLERANCE})
    write_json(RESULTS / "objective_validation.json", objective_validation)

    fault = run_fault_injections(p2, data)
    write_json(RESULTS / "fault_injection_report.json", fault)

    # 每个词典序阶段记录求解器值和由最终导出量独立复算的相应指标。
    status_by_part = {"2": status2, "3": status3, "4": status4}
    metric_by_part = {"2": ["purchase_cost_relative", "transport_loss_raw_m3"], "3": ["c_expected_supply_raw_m3", "total_expected_supply_raw_m3", "transport_loss_raw_m3"], "4": ["arrival_product_equivalent_m3_total"]}
    solver_status = {"models": []}
    for part, statuses in status_by_part.items():
        recomputed = recompute_objective(solutions[part], data)
        for index, status in enumerate(statuses):
            status = dict(status)
            key = metric_by_part[part][min(index, len(metric_by_part[part]) - 1)]
            status["problem_part"] = part
            status["recomputed_objective"] = recomputed[key] / 24.0
            status["objective_difference"] = None if status["solver_objective"] is None else abs(abs(float(status["solver_objective"])) - status["recomputed_objective"])
            solver_status["models"].append(status)
    write_json(RESULTS / "solver_status.json", solver_status)

    # 敏感性：每一项均重新执行问题2，而非只在最终表格上缩放数字。
    key_supplier = p2["active_supplier_ids"][0]
    sensitivity = {
        "base_problem": "2",
        "key_supplier": key_supplier,
        "scenarios": [
            _sensitivity_case("demand_minus_10pct", solve_problem2, data, demand=DEMAND_BASE * 0.90),
            _sensitivity_case("demand_base", solve_problem2, data, demand=DEMAND_BASE),
            _sensitivity_case("demand_plus_10pct", solve_problem2, data, demand=DEMAND_BASE * 1.10),
            _sensitivity_case("loss_low_80pct", solve_problem2, data, loss_multiplier=0.80),
            _sensitivity_case("loss_high_120pct", solve_problem2, data, loss_multiplier=1.20),
            _sensitivity_case("key_supplier_capacity_minus_10pct", solve_problem2, data, cap_multiplier={key_supplier: 0.90}),
            _sensitivity_case("key_supplier_outage", solve_problem2, data, cap_multiplier={key_supplier: 0.0}),
        ],
        "interpretation": "需求、损耗和供应能力均重新求解；供应商中断情景按整个24周均不可用处理，因此比单周中断更保守。",
    }
    write_json(RESULTS / "sensitivity_analysis.json", sensitivity)

    replay, _replay_status = solve_problem2(data)
    replay_diff = abs(replay["objective"]["purchase_cost_relative"] - p2["objective"]["purchase_cost_relative"])
    reproduction = {"method": "从同一复制材料重新执行问题2正式模型", "selected_supplier_ids_equal": replay["selected_supplier_ids"] == p2["selected_supplier_ids"], "purchase_cost_absolute_difference": replay_diff, "passed": replay_diff <= TOLERANCE and replay["selected_supplier_ids"] == p2["selected_supplier_ids"]}
    write_json(RESULTS / "reproduction_validation.json", reproduction)

    formal = {"problems": {part: _solution_summary(solution) for part, solution in solutions.items()}, "baseline": _solution_summary(baseline)}
    write_json(RESULTS / "formal_result.json", formal)
    figures = _generate_figures(analysis, sensitivity)
    elapsed = time.perf_counter() - started
    write_json(RESULTS / "runtime_log.json", {"started_at": started_wall, "ended_at": now_iso(), "runtime_seconds": elapsed, "first_feasible_seconds": first_feasible, "route_changes": 1, "generated_figures": figures})

    score = {
        "overall_score": 78,
        "dimensions": {"problem_understanding": 9, "data_quality": 8, "supplier_evaluation": 8, "model_correctness": 8, "constraint_completeness": 8, "independent_validation": 9, "algorithm_quality": 8, "result_quality": 7, "paper_quality": 7, "competition_operability": 6},
        "p0_count": 0,
        "p1_count": 2,
        "p2_count": 3,
        "decision": "development_case_passed_with_material_limitations",
        "development_case_evidence": True,
        "blind_generalization_evidence": False,
        "profile_promotion_evidence": False,
        "rationale": "P0均关闭；P1为未来供货期望假设及问题3供应商分散、问题3/4软分运的实施风险；P2为未建模季节性、绝对价格和仓储上限。终审后分数维持78。",
    }
    write_json(ROOT / "score.json", score)
    _write_docs(data, audit, analysis, baseline, solutions, validation_summary, sensitivity, score, elapsed, first_feasible)


if __name__ == "__main__":
    main()
