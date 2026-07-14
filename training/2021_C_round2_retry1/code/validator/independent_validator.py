"""不依赖求解器的结果复算与约束检查器。

本模块只读取官方附件、本轮假设、导出的原始解和输出 Excel。它不导入 ``solver``，
也不使用任何求解器目标函数或模型对象。
"""

from __future__ import annotations

import copy
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

from common.common import HORIZON, OUTPUTS, ROOT, TOLERANCE, load_official_data


def _record(
    checked_count: int,
    violations: list[dict[str, Any]],
    threshold: float = TOLERANCE,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload = {
        "checked_count": int(checked_count),
        "violation_count": len(violations),
        "max_violation": float(max((item["magnitude"] for item in violations), default=0.0)),
        "threshold": threshold,
        "examples": violations[:5],
        "passed": len(violations) == 0,
    }
    if extra:
        payload.update(extra)
    return payload


def _arrays(solution: dict[str, Any]) -> dict[str, np.ndarray]:
    """把导出的 JSON 数组转换为数组；形状错误由各项检查报出。"""
    return {
        "orders": np.asarray(solution["orders_raw_m3"], dtype=float),
        "supply": np.asarray(solution["expected_supply_raw_m3"], dtype=float),
        "shipments": np.asarray(solution["shipments_raw_m3"], dtype=float),
        "arrival_raw": np.asarray(solution["arrivals_raw_m3"], dtype=float),
        "arrival_product": np.asarray(solution["arrivals_product_equivalent_m3"], dtype=float),
        "losses": np.asarray(solution["losses_by_transporter_raw_m3"], dtype=float),
        "inventory": np.asarray(solution["inventory_product_equivalent_m3"], dtype=float),
    }


def recompute_supplier_metrics(data: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    """从官方附件重新计算供应商基础指标，供问题1和复核共用。"""
    return (data or load_official_data())["supplier_metrics"]


def recompute_objective(solution: dict[str, Any], data: dict[str, Any] | None = None) -> dict[str, float]:
    """仅基于导出量、类别和损耗率复算成本、损耗和到货能力。"""
    data = data or load_official_data()
    arrays = _arrays(solution)
    supply = arrays["supply"]
    shipments = arrays["shipments"]
    return {
        "purchase_cost_relative": float(np.sum(supply * data["unit_cost"][:, None])),
        "transport_loss_raw_m3": float(np.sum(shipments * data["loss_mean"][None, None, :])),
        "total_expected_supply_raw_m3": float(supply.sum()),
        "a_expected_supply_raw_m3": float(supply[np.array(data["material_types"]) == "A"].sum()),
        "c_expected_supply_raw_m3": float(supply[np.array(data["material_types"]) == "C"].sum()),
        "arrival_product_equivalent_m3_total": float(arrays["arrival_product"].sum()),
    }


def check_supplier_selection(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    known = set(data["supplier_ids"])
    selected = solution.get("selected_supplier_ids", [])
    violations: list[dict[str, Any]] = []
    for supplier_id in selected:
        if supplier_id not in known:
            violations.append({"supplier_id": supplier_id, "magnitude": 1.0, "reason": "未知供应商"})
    if len(selected) != len(set(selected)):
        violations.append({"magnitude": 1.0, "reason": "供应商选择列表存在重复"})
    selected_set = set(selected)
    for index, supplier_id in enumerate(data["supplier_ids"]):
        if arrays["orders"][index].max(initial=0.0) > TOLERANCE and supplier_id not in selected_set:
            violations.append({"supplier_id": supplier_id, "magnitude": 1.0, "reason": "未选择供应商存在订货"})
    if solution.get("problem_part") == "2":
        required = int(solution.get("selection_minimum_count", len(selected)))
        if "baseline_selected_count" in solution:
            if len(selected) < required:
                violations.append({"magnitude": abs(len(selected) - required), "reason": "基线供应商数低于最小能力数"})
        elif len(selected) != required:
            violations.append({"magnitude": abs(len(selected) - required), "reason": "最少供应商数量不一致"})
    return _record(len(selected) + len(data["supplier_ids"]), violations)


def check_order_nonnegative(solution: dict[str, Any], _data: dict[str, Any]) -> dict[str, Any]:
    orders = _arrays(solution)["orders"]
    bad = np.argwhere(orders < -TOLERANCE)
    violations = [
        {"supplier_index": int(i), "week": int(t + 1), "magnitude": float(-orders[i, t]), "reason": "订货量为负"}
        for i, t in bad[:5]
    ]
    return _record(orders.size, violations)


def check_supplier_capacity(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    supply = _arrays(solution)["supply"]
    excess = supply - data["regular_order_capacity"][:, None]
    bad = np.argwhere(excess > TOLERANCE)
    violations = [
        {"supplier_id": data["supplier_ids"][i], "week": int(t + 1), "magnitude": float(excess[i, t]), "reason": "超过常规供货能力"}
        for i, t in bad[:5]
    ]
    return _record(excess.size, violations)


def check_material_conversion(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    recomputed = (arrays["arrival_raw"] / data["raw_per_product"][:, None]).sum(axis=0)
    difference = np.abs(recomputed - arrays["arrival_product"])
    violations = [
        {"week": int(t + 1), "magnitude": float(difference[t]), "reason": "原料到货与产品等价量不一致"}
        for t in np.flatnonzero(difference > TOLERANCE)[:5]
    ]
    return _record(HORIZON, violations)


def check_weekly_production_requirement(solution: dict[str, Any], _data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    demand = float(solution["demand_product_m3_per_week"])
    available = arrays["inventory"][:-1] + arrays["arrival_product"] - arrays["inventory"][1:]
    shortfall = demand - available
    violations = [
        {"week": int(t + 1), "magnitude": float(shortfall[t]), "reason": "可用原料不足生产需求"}
        for t in np.flatnonzero(shortfall > TOLERANCE)[:5]
    ]
    return _record(HORIZON, violations)


def check_inventory_balance(solution: dict[str, Any], _data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    demand = float(solution["demand_product_m3_per_week"])
    residual = arrays["inventory"][1:] - (arrays["inventory"][:-1] + arrays["arrival_product"] - demand)
    violations = [
        {"week": int(t + 1), "magnitude": float(abs(residual[t])), "reason": "库存守恒残差"}
        for t in np.flatnonzero(np.abs(residual) > TOLERANCE)[:5]
    ]
    return _record(HORIZON, violations)


def check_initial_inventory(solution: dict[str, Any], _data: dict[str, Any]) -> dict[str, Any]:
    inventory = _arrays(solution)["inventory"]
    demand = float(solution["demand_product_m3_per_week"])
    difference = abs(inventory[0] - 2.0 * demand)
    return _record(1, [] if difference <= TOLERANCE else [{"magnitude": float(difference), "reason": "初始库存不是两周需求"}])


def check_terminal_inventory(solution: dict[str, Any], _data: dict[str, Any]) -> dict[str, Any]:
    inventory = _arrays(solution)["inventory"]
    demand = float(solution["demand_product_m3_per_week"])
    shortfall = 2.0 * demand - inventory[-1]
    return _record(1, [] if shortfall <= TOLERANCE else [{"magnitude": float(shortfall), "reason": "期末库存低于两周安全库存"}])


def check_transporter_capacity(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    shipments = _arrays(solution)["shipments"]
    loads = shipments.sum(axis=0)
    excess = loads - 6_000.0
    bad = np.argwhere(excess > TOLERANCE)
    violations = [
        {"week": int(t + 1), "transporter_id": data["transporter_ids"][j], "magnitude": float(excess[t, j]), "reason": "超过周承运能力"}
        for t, j in bad[:5]
    ]
    return _record(loads.size, violations)


def check_supplier_transporter_assignment(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    shipments = _arrays(solution)["shipments"]
    split = (shipments > TOLERANCE).sum(axis=2)
    hard = bool(solution.get("model_metadata", {}).get("single_carrier_hard", False))
    bad = np.argwhere(split > 1)
    violations = [
        {"supplier_id": data["supplier_ids"][i], "week": int(t + 1), "magnitude": float(split[i, t] - 1), "reason": "同一供应商被多个转运商分运"}
        for i, t in bad[:5]
    ] if hard else []
    return _record(
        split.size,
        violations,
        extra={"constraint_kind": "hard" if hard else "soft", "soft_split_occurrences": int((split > 1).sum())},
    )


def check_transport_loss(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    expected = (arrays["shipments"] * data["loss_mean"][None, None, :]).sum(axis=0)
    difference = np.abs(expected - arrays["losses"])
    bad = np.argwhere(difference > TOLERANCE)
    violations = [
        {"week": int(t + 1), "transporter_id": data["transporter_ids"][j], "magnitude": float(difference[t, j]), "reason": "损耗量与损耗率不一致"}
        for t, j in bad[:5]
    ]
    return _record(difference.size, violations)


def check_arrival_quantity(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    expected = (arrays["shipments"] * (1.0 - data["loss_mean"])[None, None, :]).sum(axis=2)
    difference = np.abs(expected - arrays["arrival_raw"])
    bad = np.argwhere(difference > TOLERANCE)
    violations = [
        {"supplier_id": data["supplier_ids"][i], "week": int(t + 1), "magnitude": float(difference[i, t]), "reason": "接收量与转运量、损耗不一致"}
        for i, t in bad[:5]
    ]
    return _record(difference.size, violations)


def check_order_transport_consistency(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    difference = np.abs(arrays["supply"] - arrays["shipments"].sum(axis=2))
    bad = np.argwhere(difference > TOLERANCE)
    violations = [
        {"supplier_id": data["supplier_ids"][i], "week": int(t + 1), "magnitude": float(difference[i, t]), "reason": "预计供货量与转运量不一致"}
        for i, t in bad[:5]
    ]
    return _record(difference.size, violations)


def check_units_and_aggregation(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    arrays = _arrays(solution)
    violations: list[dict[str, Any]] = []
    if not np.all(np.isin(data["raw_per_product"], [0.60, 0.66, 0.72])):
        violations.append({"magnitude": 1.0, "reason": "材料转换系数与题面不一致"})
    if np.any(data["loss_mean"] < 0) or np.any(data["loss_mean"] >= 1):
        violations.append({"magnitude": 1.0, "reason": "损耗率未按比例口径处理"})
    if arrays["shipments"].shape[2] != len(data["transporter_ids"]):
        violations.append({"magnitude": 1.0, "reason": "转运商聚合维度错误"})
    return _record(3, violations)


def _excel_number(value: Any) -> float:
    return 0.0 if value is None else float(value)


def check_excel_output_consistency(solution: dict[str, Any], data: dict[str, Any], output_dir: Path = OUTPUTS) -> dict[str, Any]:
    """回读附件A/B，并逐格比对原始决策变量与模板行列映射。"""
    part = solution["problem_part"]
    order_path = output_dir / "附件A 订购方案数据结果.xlsx"
    transport_path = output_dir / "附件B 转运方案数据结果.xlsx"
    if not order_path.exists() or not transport_path.exists():
        return _record(0, [{"magnitude": 1.0, "reason": "未找到输出 Excel"}])
    orders = _arrays(solution)["orders"]
    shipments = _arrays(solution)["shipments"]
    order_book = load_workbook(order_path, data_only=False, read_only=True)
    transport_book = load_workbook(transport_path, data_only=False, read_only=True)
    order_sheet = order_book[f"问题{part}的订购方案结果"]
    transport_sheet = transport_book[f"问题{part}的转运方案结果"]
    id_to_index = {supplier_id: i for i, supplier_id in enumerate(data["supplier_ids"])}
    violations: list[dict[str, Any]] = []
    checked = 0
    order_rows = order_sheet.iter_rows(min_row=7, max_row=408, min_col=1, max_col=25, values_only=True)
    transport_rows = transport_sheet.iter_rows(min_row=7, max_row=408, min_col=1, max_col=193, values_only=True)
    for row, (order_values, transport_values) in enumerate(zip(order_rows, transport_rows), start=7):
        supplier_id = str(order_values[0])
        index = id_to_index[supplier_id]
        for week in range(HORIZON):
            actual = _excel_number(order_values[week + 1])
            expected = float(orders[index, week])
            checked += 1
            if abs(actual - expected) > TOLERANCE:
                violations.append({"cell": f"R{row}C{week + 2}", "magnitude": abs(actual - expected), "reason": "订购模板与原始方案不一致"})
        for week in range(HORIZON):
            for transporter in range(8):
                column = 2 + week * 8 + transporter
                actual = _excel_number(transport_values[column - 1])
                expected = float(shipments[index, week, transporter])
                checked += 1
                if abs(actual - expected) > TOLERANCE:
                    violations.append({"cell": f"R{row}C{column}", "magnitude": abs(actual - expected), "reason": "转运模板与原始方案不一致"})
    return _record(checked, violations)


def check_all_constraints(solution: dict[str, Any], data: dict[str, Any] | None = None, output_dir: Path = OUTPUTS) -> dict[str, Any]:
    """逐类输出硬约束报告；不以单个布尔值掩盖细项。"""
    data = data or load_official_data()
    checks = {
        "supplier_selection": check_supplier_selection(solution, data),
        "supplier_capacity": check_supplier_capacity(solution, data),
        "order_nonnegative": check_order_nonnegative(solution, data),
        "material_conversion": check_material_conversion(solution, data),
        "production_requirement": check_weekly_production_requirement(solution, data),
        "inventory_balance": check_inventory_balance(solution, data),
        "initial_inventory": check_initial_inventory(solution, data),
        "terminal_inventory": check_terminal_inventory(solution, data),
        "transporter_capacity": check_transporter_capacity(solution, data),
        "transporter_assignment": check_supplier_transporter_assignment(solution, data),
        "transport_loss": check_transport_loss(solution, data),
        "arrival_quantity": check_arrival_quantity(solution, data),
        "order_transport_consistency": check_order_transport_consistency(solution, data),
        "output_template_consistency": check_excel_output_consistency(solution, data, output_dir),
        "units_and_aggregation": check_units_and_aggregation(solution, data),
    }
    total = sum(item["violation_count"] for item in checks.values())
    return {"checks": checks, "total_hard_violations": total, "passed": total == 0}


def _has_detection(report: dict[str, Any], expected: str) -> bool:
    return report["checks"].get(expected, {}).get("violation_count", 0) > 0


def run_fault_injections(solution: dict[str, Any], data: dict[str, Any] | None = None) -> dict[str, Any]:
    """对问题2解进行不少于12项的定向篡改，确认检查器不会静默放过错误。"""
    data = data or load_official_data()
    arrays = _arrays(solution)
    active_supplier = int(np.flatnonzero(arrays["supply"].sum(axis=1) > TOLERANCE)[0])
    active_week = 0
    active_transporter = int(np.flatnonzero(arrays["shipments"][active_supplier, active_week] > TOLERANCE)[0])
    alternative_transporter = (active_transporter + 1) % len(data["transporter_ids"])
    selected_set = set(solution["selected_supplier_ids"])
    unselected = next(index for index, supplier_id in enumerate(data["supplier_ids"]) if supplier_id not in selected_set)

    def changed(mutator: Any, expected: str) -> dict[str, Any]:
        trial = copy.deepcopy(solution)
        mutator(trial)
        report = check_all_constraints(trial, data)
        return {"expected_detector": expected, "detected": _has_detection(report, expected), "detected_categories": [key for key, value in report["checks"].items() if value["violation_count"] > 0]}

    tests: list[dict[str, Any]] = []
    tests.append({"id": 1, "name": "负订货量", **changed(lambda s: s["orders_raw_m3"][active_supplier].__setitem__(0, -1.0), "order_nonnegative")})
    tests.append({"id": 2, "name": "超过供应能力", **changed(lambda s: s["expected_supply_raw_m3"][active_supplier].__setitem__(0, float(data["regular_order_capacity"][active_supplier] + 1.0)), "supplier_capacity")})
    tests.append({"id": 3, "name": "生产原料不足", **changed(lambda s: s["inventory_product_equivalent_m3"].__setitem__(1, 100000.0), "production_requirement")})
    tests.append({"id": 4, "name": "库存守恒篡改", **changed(lambda s: s["inventory_product_equivalent_m3"].__setitem__(1, s["inventory_product_equivalent_m3"][1] + 10.0), "inventory_balance")})
    tests.append({"id": 5, "name": "初始库存错误", **changed(lambda s: s["inventory_product_equivalent_m3"].__setitem__(0, 0.0), "initial_inventory")})
    tests.append({"id": 6, "name": "转运商超载", **changed(lambda s: s["shipments_raw_m3"][active_supplier][0].__setitem__(active_transporter, 7000.0), "transporter_capacity")})
    tests.append({"id": 7, "name": "供应商多转运商分运", **changed(lambda s: s["shipments_raw_m3"][active_supplier][0].__setitem__(alternative_transporter, 1.0), "transporter_assignment")})
    tests.append({"id": 8, "name": "损耗计算错误", **changed(lambda s: s["losses_by_transporter_raw_m3"][0].__setitem__(active_transporter, s["losses_by_transporter_raw_m3"][0][active_transporter] + 1.0), "transport_loss")})
    tests.append({"id": 9, "name": "到货量不一致", **changed(lambda s: s["arrivals_raw_m3"][active_supplier].__setitem__(0, s["arrivals_raw_m3"][active_supplier][0] + 1.0), "arrival_quantity")})
    tests.append({"id": 10, "name": "订购转运不一致", **changed(lambda s: s["expected_supply_raw_m3"][active_supplier].__setitem__(0, s["expected_supply_raw_m3"][active_supplier][0] + 1.0), "order_transport_consistency")})
    objective_trial = copy.deepcopy(solution)
    objective_trial["objective"]["purchase_cost_relative"] += 1.0
    objective_ok = abs(objective_trial["objective"]["purchase_cost_relative"] - recompute_objective(objective_trial, data)["purchase_cost_relative"]) > TOLERANCE
    tests.append({"id": 11, "name": "人为篡改目标值", "expected_detector": "objective_recompute", "detected": objective_ok, "detected_categories": ["objective_recompute"] if objective_ok else []})
    tests.append({"id": 12, "name": "Excel与原始变量不一致", **changed(lambda s: s["orders_raw_m3"][active_supplier].__setitem__(0, s["orders_raw_m3"][active_supplier][0] + 1.0), "output_template_consistency")})
    tests.append({"id": 13, "name": "未选择供应商订货", **changed(lambda s: s["orders_raw_m3"][unselected].__setitem__(0, 1.0), "supplier_selection")})
    tests.append({"id": 14, "name": "转换系数影响未同步", **changed(lambda s: s["arrivals_product_equivalent_m3"].__setitem__(0, s["arrivals_product_equivalent_m3"][0] + 1.0), "material_conversion")})
    passed = sum(bool(item["detected"]) for item in tests)
    return {"fault_injection_total": len(tests), "fault_injection_passed": passed, "fault_injection_pass_rate": passed / len(tests), "tests": tests}
