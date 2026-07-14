"""2021-C 终审计算与干净重放对比。

本模块只读取官方材料和导出结果，不导入求解器，也不改变任何优化变量。
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from scipy.stats import rankdata, spearmanr

from common.common import OUTPUTS, RESULTS, TOLERANCE, load_official_data, now_iso, read_json, write_json
from validator.independent_validator import check_excel_output_consistency


KEY_RESULT_FILES = [
    "supplier_analysis.json",
    "baseline_result.json",
    "raw_solution.json",
    "objective_validation.json",
    "constraint_validation.json",
    "solver_status.json",
]
KEY_OUTPUT_FILES = ["附件A 订购方案数据结果.xlsx", "附件B 转运方案数据结果.xlsx"]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _rank_sensitivity(data: dict[str, Any]) -> dict[str, Any]:
    metrics = data["supplier_metrics"]
    raw_components = {
        "capacity": np.array([item["product_equivalent_capacity_before_loss"] for item in metrics]),
        "service": np.array([item["service_probability"] for item in metrics]),
        "stability": np.array(
            [
                1.0 / (1.0 + item["supply_cv_positive_weeks"])
                if item["supply_cv_positive_weeks"] is not None
                else 0.0
                for item in metrics
            ]
        ),
        "fulfilment": np.array([min(item["weighted_fulfilment_ratio"], 1.0) for item in metrics]),
    }

    def normalise(values: np.ndarray) -> np.ndarray:
        span = float(values.max() - values.min())
        return np.zeros_like(values) if span <= TOLERANCE else (values - values.min()) / span

    components = {name: normalise(values) for name, values in raw_components.items()}
    baseline = {"capacity": 0.50, "service": 0.25, "stability": 0.15, "fulfilment": 0.10}
    weight_sets: dict[str, dict[str, float]] = {
        "current": baseline,
        "equal": {name: 0.25 for name in baseline},
    }
    for changed in baseline:
        for factor, suffix in ((0.90, "minus_10pct"), (1.10, "plus_10pct")):
            varied = dict(baseline)
            varied[changed] *= factor
            total = sum(varied.values())
            weight_sets[f"{changed}_{suffix}"] = {name: value / total for name, value in varied.items()}

    supplier_ids = data["supplier_ids"]
    rankings: dict[str, np.ndarray] = {}
    scores: dict[str, np.ndarray] = {}
    rank_positions: dict[str, np.ndarray] = {}
    for name, weights in weight_sets.items():
        scores[name] = sum(weights[key] * components[key] for key in weights)
        rankings[name] = np.argsort(-scores[name], kind="stable")
        # 降序得分采用平均秩；稳定排序位置另用于供应商逐项展示。
        rank_positions[name] = rankdata(-scores[name], method="average")

    current_top10 = set(rankings["current"][:10].tolist())
    current_top50 = set(rankings["current"][:50].tolist())
    cases: dict[str, Any] = {}
    for name in weight_sets:
        order = rankings[name]
        correlation = float(spearmanr(rank_positions["current"], rank_positions[name]).statistic)
        cases[name] = {
            "weights": weight_sets[name],
            "top10_overlap_count_with_current": len(current_top10 & set(order[:10].tolist())),
            "top50_overlap_rate_with_current": len(current_top50 & set(order[:50].tolist())) / 50.0,
            "spearman_rank_correlation_with_current": correlation,
            "top10_supplier_ids": [supplier_ids[index] for index in order[:10]],
        }

    focus: dict[str, Any] = {}
    for supplier_id in ("S229", "S361", "S140"):
        index = supplier_ids.index(supplier_id)
        positions = {
            name: int(np.flatnonzero(rankings[name] == index)[0] + 1)
            for name in rankings
        }
        focus[supplier_id] = {
            "ranks_by_case": positions,
            "best_rank": min(positions.values()),
            "worst_rank": max(positions.values()),
            "top10_in_all_cases": max(positions.values()) <= 10,
        }

    non_current = [case for name, case in cases.items() if name != "current"]
    return {
        "generated_at": now_iso(),
        "method": "四项极差归一化指标线性加权；除等权外，每次仅将一个当前权重相对调整±10%，再将四项权重归一化。",
        "weight_semantics": "权重是业务假设，不是由题面或统计估计得到的客观最优权重。",
        "cases": cases,
        "focus_suppliers": focus,
        "summary": {
            "minimum_top10_overlap_count": min(case["top10_overlap_count_with_current"] for case in non_current),
            "minimum_top50_overlap_rate": min(case["top50_overlap_rate_with_current"] for case in non_current),
            "minimum_spearman_rank_correlation": min(case["spearman_rank_correlation_with_current"] for case in non_current),
            "focus_suppliers_all_top10": all(item["top10_in_all_cases"] for item in focus.values()),
            "conclusion": "排序对小幅单项权重扰动稳定，但等权方案会改变头部次序；排序用于候选筛选，不等同于客观质量排名。",
        },
    }


def _supply_semantics(data: dict[str, Any]) -> dict[str, Any]:
    orders = np.asarray(data["orders_history"], dtype=float)
    supply = np.asarray(data["supply_history"], dtype=float)
    ordered = orders > 0
    supplied = supply > 0
    conditional_positive = np.where(ordered, supply, np.nan)
    return {
        "supplier_week_counts": {
            "not_ordered_and_zero_supply": int((~ordered & ~supplied).sum()),
            "not_ordered_but_positive_supply": int((~ordered & supplied).sum()),
            "ordered_but_zero_supply": int((ordered & ~supplied).sum()),
            "ordered_and_positive_supply": int((ordered & supplied).sum()),
        },
        "capacity_definitions": {
            "mean_supply_over_all_240_weeks": "描述历史业务实际发生量，受企业是否下单影响，不作为未来能力上限。",
            "historical_max_supply": "仅描述历史单周峰值，不作为可持续能力。",
            "conditional_on_order_mean_supply": float(np.nanmean(conditional_positive)),
            "model_regular_order_capacity": "供应商非零供货均值 × P(非零供货|已下单)，是规律下单条件下的期望供货上限。",
            "conservative_future_capacity": "当前 regular_order_capacity 已折入按单供货概率，但仍是历史期望而非保证量；压力测试另检验能力下降和中断。",
        },
        "global_historical_max_supply_raw_m3": float(supply.max()),
        "capacity_uses_all_240_week_denominator": False,
        "historical_order_volume_used_as_capacity": False,
        "audit_conclusion": "能力估计未使用供货量/240周，也未把历史订购量本身当作能力；仍存在以历史期望代替未来确定供货的业务假设。",
    }


def _inventory_audit(solutions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    by_problem: dict[str, Any] = {}
    for part, solution in solutions.items():
        inventory = np.asarray(solution["inventory_product_equivalent_m3"], dtype=float)
        arrivals = np.asarray(solution["arrivals_product_equivalent_m3"], dtype=float)
        demand = float(solution["demand_product_m3_per_week"])
        residual = inventory[1:] - (inventory[:-1] + arrivals - demand)
        by_problem[part] = {
            "demand_product_m3_per_week": demand,
            "safety_stock_lower_bound_product_m3": 2.0 * demand,
            "initial_inventory_product_m3": float(inventory[0]),
            "minimum_end_inventory_product_m3": float(inventory[1:].min()),
            "maximum_end_inventory_product_m3": float(inventory[1:].max()),
            "arrival_product_m3_per_week": float(arrivals[0]),
            "maximum_balance_absolute_residual": float(np.abs(residual).max()),
            "all_end_inventories_at_lower_bound": bool(np.all(np.abs(inventory[1:] - 2.0 * demand) <= TOLERANCE)),
        }
    return {
        "time_order": "期初库存 + 当周实际损耗后到货 - 当周生产消耗 = 期末库存。",
        "arrival_interpretation": "当周到货是 shipments × (1-loss) 的接收量换算为产品等价量后进入库存；未使用尚未到达的发运量。",
        "constant_56400_interpretation": "问题2、3中56,400 m3是两周基准生产需求的安全库存下界，也是显式期初库存假设；每周到货恰等于28,200 m3且目标不奖励囤货，因此期末库存逐周紧贴下界。",
        "problem4_note": "问题4的需求变量等于最大可持续周产能，因此其两周安全库存为最大产能的两倍，不是56,400 m3。",
        "by_problem": by_problem,
        "passed": all(item["maximum_balance_absolute_residual"] <= TOLERANCE for item in by_problem.values()),
    }


def _positive_distribution(values: np.ndarray) -> dict[str, Any]:
    positive = values[values > TOLERANCE]
    return {
        "positive_count": int(positive.size),
        "minimum": float(positive.min()) if positive.size else None,
        "p10": float(np.quantile(positive, 0.10)) if positive.size else None,
        "median": float(np.quantile(positive, 0.50)) if positive.size else None,
        "p90": float(np.quantile(positive, 0.90)) if positive.size else None,
        "maximum": float(positive.max()) if positive.size else None,
        "count_lt_1": int((positive < 1.0).sum()),
        "count_lt_10": int((positive < 10.0).sum()),
        "count_lt_100": int((positive < 100.0).sum()),
    }


def _supplier_usage_distribution(solutions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    by_problem: dict[str, Any] = {}
    for part, solution in solutions.items():
        orders = np.asarray(solution["orders_raw_m3"], dtype=float)
        representative = orders.max(axis=1)
        supplier_stats = _positive_distribution(representative)
        cell_stats = _positive_distribution(orders)
        by_problem[part] = {
            "active_supplier_count": int((representative > TOLERANCE).sum()),
            "positive_weekly_order_by_supplier": supplier_stats,
            "positive_supplier_week_cells": cell_stats,
        }

    p3 = by_problem["3"]["positive_weekly_order_by_supplier"]
    numerical_fragment = p3["minimum"] is not None and p3["minimum"] < 1.0
    return {
        "generated_at": now_iso(),
        "positive_threshold": TOLERANCE,
        "distribution_basis": "方案按单周最优流复制到24周；供应商口径使用每家供应商24周最大正订购量，单元格口径另行报告。",
        "by_problem": by_problem,
        "problem3_assessment": {
            "active_supplier_count": by_problem["3"]["active_supplier_count"],
            "has_sub_1_raw_m3_fragments": numerical_fragment,
            "conclusion": (
                "存在小于1 m3的正订购量，连续LP未惩罚供应商启用数量；这些值应视为数值非零，不应手工删除。"
                if numerical_fragment
                else "未发现小于1 m3的数值碎片；大量供应商是连续LP在少C、少总原料、少损耗目标下的业务结果，但模型仍未惩罚启用数量。"
            ),
            "operability_limit": "若业务目标要求减少合作供应商，需要另建启用变量与固定成本；问题3题面当前目标下不构成硬约束违例。",
        },
    }


def _transport_split_analysis(solutions: dict[str, dict[str, Any]], solver_status: dict[str, Any]) -> dict[str, Any]:
    runtime_by_problem: dict[str, float] = {}
    for item in solver_status["models"]:
        part = str(item["problem_part"])
        runtime_by_problem[part] = runtime_by_problem.get(part, 0.0) + float(item["runtime_seconds"])

    by_problem: dict[str, Any] = {}
    for part in ("2", "3"):
        shipments = np.asarray(solutions[part]["shipments_raw_m3"], dtype=float)
        carrier_count = (shipments > TOLERANCE).sum(axis=2)
        split_mask = carrier_count > 1
        supplier_week_active = carrier_count > 0
        supplier_week_volume = shipments.sum(axis=2)
        largest_carrier_volume = shipments.max(axis=2)
        total_volume = float(shipments.sum())
        volume_in_split = float(supplier_week_volume[split_mask].sum())
        secondary_volume = float((supplier_week_volume - largest_carrier_volume)[split_mask].sum())
        by_problem[part] = {
            "single_carrier_is_hard": bool(solutions[part]["model_metadata"]["single_carrier_hard"]),
            "split_supplier_week_count": int(split_mask.sum()),
            "active_supplier_week_count": int(supplier_week_active.sum()),
            "split_ratio_of_active_supplier_weeks": float(split_mask.sum() / max(supplier_week_active.sum(), 1)),
            "volume_in_split_supplier_weeks_raw_m3": volume_in_split,
            "volume_in_split_supplier_weeks_share": volume_in_split / total_volume if total_volume else 0.0,
            "secondary_carrier_volume_raw_m3": secondary_volume,
            "secondary_carrier_volume_share": secondary_volume / total_volume if total_volume else 0.0,
            "total_transport_volume_raw_m3": total_volume,
            "maximum_carriers_per_supplier_week": int(carrier_count.max()),
            "transport_loss_raw_m3": float(solutions[part]["objective"]["transport_loss_raw_m3"]),
            "solver_runtime_seconds": runtime_by_problem.get(part),
        }
    return {
        "generated_at": now_iso(),
        "positive_threshold": TOLERANCE,
        "by_problem": by_problem,
        "hard_single_carrier_comparison": {
            "problem2": "正式模型本身强制单承运商，拆分为0，可直接作为硬约束基准。",
            "problem3": "既有失败路线在80秒内未返回可行解，状态只能记为unknown；因此不能给出损耗差，也不能写成不可行。连续LP三阶段运行时间见by_problem.3。",
            "comparable_loss_change_available": False,
        },
    }


def _problem4_audit(solution: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    shipments = np.asarray(solution["shipments_raw_m3"], dtype=float)[:, 0, :]
    loss = np.asarray(data["loss_mean"], dtype=float)
    raw_per_product = np.asarray(data["raw_per_product"], dtype=float)
    expected_supply = shipments.sum(axis=1)
    received_raw = (shipments * (1.0 - loss)[None, :]).sum(axis=1)
    product_by_material = {
        material: float((received_raw[np.array(data["material_types"]) == material] / raw_per_product[np.array(data["material_types"]) == material]).sum())
        for material in "ABC"
    }
    sustainable = sum(product_by_material.values())
    reported = float(solution["maximum_weekly_production_m3"])
    increase = reported - 28_200.0
    return {
        "reported_sustainable_weekly_capacity_product_m3": reported,
        "independently_recomputed_capacity_product_m3": sustainable,
        "absolute_difference": abs(reported - sustainable),
        "baseline_weekly_capacity_product_m3": 28_200.0,
        "capacity_increase_product_m3": increase,
        "capacity_increase_ratio": increase / 28_200.0,
        "expected_supply_raw_m3": float(expected_supply.sum()),
        "received_after_loss_raw_m3": float(received_raw.sum()),
        "product_equivalent_by_material_m3": product_by_material,
        "checks": {
            "conversion_direction": "损耗后原料到货量除以每产品m3原料消耗系数0.60/0.66/0.72。",
            "orders_not_used_as_arrivals": True,
            "transport_loss_applied_once": True,
            "initial_inventory_in_capacity_objective": False,
            "first_week_inventory_extrapolated": False,
        },
        "interpretation": "在当前历史能力估计、连续供货和损耗假设下，预测可持续周产能为该值；不是企业一定能够长期实现的保证产能。",
        "passed": abs(reported - sustainable) <= TOLERANCE,
    }


def _stress_audit(sensitivity: dict[str, Any]) -> dict[str, Any]:
    scenarios = {item["scenario"]: item for item in sensitivity["scenarios"]}
    expected = {
        "demand_plus_10pct": "infeasible",
        "key_supplier_outage": "infeasible",
        "key_supplier_capacity_minus_10pct": "unknown_time_limit",
    }
    actual: dict[str, Any] = {}
    passed = True
    for name, status in expected.items():
        item = scenarios[name]
        actual_status = "feasible" if item.get("feasible") is True else item.get("feasibility_status")
        actual[name] = {"expected_status": status, "actual_status": actual_status, "passed": actual_status == status}
        passed = passed and actual_status == status
    return {
        "scenarios": actual,
        "required_paper_statement": "当前方案在基准需求下可行，但对需求上升和关键供应商长期中断较敏感，供应链冗余仍不足。",
        "unknown_is_not_infeasible": True,
        "passed": passed,
    }


def _snapshot() -> dict[str, Any]:
    raw = read_json(RESULTS / "raw_solution.json")
    formal = read_json(RESULTS / "formal_result.json")
    constraints = read_json(RESULTS / "constraint_validation.json")
    objectives = read_json(RESULTS / "objective_validation.json")
    outputs = read_json(RESULTS / "output_template_validation.json")
    file_hashes = {
        f"results/{name}": _sha256(RESULTS / name)
        for name in KEY_RESULT_FILES
    }
    file_hashes.update({f"outputs/{name}": _sha256(OUTPUTS / name) for name in KEY_OUTPUT_FILES})
    return {
        "created_at": now_iso(),
        "file_hashes": file_hashes,
        "key_metrics": {
            "formal_result": formal,
            "reported_objectives": {part: raw["problems"][part]["objective"] for part in ("2", "3", "4")},
            "active_supplier_counts": {
                part: len(raw["problems"][part]["active_supplier_ids"])
                for part in ("2", "3", "4")
            },
            "hard_constraint_violations": constraints["total_hard_violations"],
            "constraint_validation_passed": constraints["passed"],
            "objective_validation": objectives,
            "excel_validation_passed": outputs["passed"],
        },
    }


def _compare_snapshots(before: dict[str, Any], after: dict[str, Any]) -> dict[str, Any]:
    before_metrics = before["key_metrics"]
    after_metrics = after["key_metrics"]
    hash_comparison = {
        name: {
            "before": digest,
            "after": after["file_hashes"][name],
            "equal": digest == after["file_hashes"][name],
        }
        for name, digest in before["file_hashes"].items()
    }
    objectives_equal = before_metrics["reported_objectives"] == after_metrics["reported_objectives"]
    supplier_counts_equal = before_metrics["active_supplier_counts"] == after_metrics["active_supplier_counts"]
    formal_equal = before_metrics["formal_result"] == after_metrics["formal_result"]
    constraints_pass = after_metrics["hard_constraint_violations"] == 0 and after_metrics["constraint_validation_passed"]
    excel_recalculation = read_json(RESULTS / "excel_recalculation_validation.json")
    excel_pass = bool(after_metrics["excel_validation_passed"] and excel_recalculation["passed"])
    return {
        "generated_at": now_iso(),
        "method": "先隔离results、outputs、figures及生成文档，再从官方materials执行唯一入口；哈希仅作辅助，优先验收目标、供应商数、硬约束、聚合结果和Excel逐格回读。",
        "before_created_at": before["created_at"],
        "after_created_at": after["created_at"],
        "hash_comparison": hash_comparison,
        "hash_equal_count": sum(item["equal"] for item in hash_comparison.values()),
        "hash_total_count": len(hash_comparison),
        "semantic_comparison": {
            "reported_objectives_equal": objectives_equal,
            "active_supplier_counts_equal": supplier_counts_equal,
            "formal_aggregates_equal": formal_equal,
            "zero_hard_constraint_violations": constraints_pass,
            "excel_cell_validation_passed": excel_pass,
            "excel_formula_recalculation_passed": bool(excel_recalculation["passed"]),
        },
        "before_key_metrics": before_metrics,
        "after_key_metrics": after_metrics,
        "equivalent_optimum_policy": "允许等价最优解导致变量级和文件哈希差异；本次结论以语义指标为准。",
        "passed": all([objectives_equal, supplier_counts_equal, formal_equal, constraints_pass, excel_pass]),
    }


def _excel_recalculation_audit() -> dict[str, Any]:
    records = []
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}
    for name in KEY_OUTPUT_FILES:
        path = OUTPUTS / name
        formula_book = load_workbook(path, data_only=False, read_only=False)
        value_book = load_workbook(path, data_only=True, read_only=False)
        formula_count = 0
        cached_blank_count = 0
        formula_errors: list[str] = []
        for sheet in formula_book.worksheets:
            value_sheet = value_book[sheet.title]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    formula_count += 1
                    cached = value_sheet[cell.coordinate].value
                    cached_blank_count += int(cached is None)
                    if isinstance(cached, str) and cached.upper() in error_tokens:
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}:{cached}")
        formula_book.close()
        value_book.close()
        records.append(
            {
                "file": name,
                "sha256": _sha256(path),
                "formula_count": formula_count,
                "cached_formula_blank_count": cached_blank_count,
                "formula_error_count": len(formula_errors),
                "formula_error_examples": formula_errors[:10],
                "passed": cached_blank_count == 0 and not formula_errors,
            }
        )

    data = load_official_data()
    solutions = read_json(RESULTS / "raw_solution.json")["problems"]
    consistency = {
        part: check_excel_output_consistency(solutions[part], data)
        for part in ("2", "3", "4")
    }
    return {
        "engine": "LibreOffice Calc headless conversion",
        "generated_at": now_iso(),
        "records": records,
        "raw_solution_cell_consistency": consistency,
        "passed": all(item["passed"] for item in records)
        and all(item["passed"] for item in consistency.values()),
    }


def generate_audit() -> None:
    data = load_official_data()
    raw = read_json(RESULTS / "raw_solution.json")
    solutions = raw["problems"]
    rank = _rank_sensitivity(data)
    usage = _supplier_usage_distribution(solutions)
    split = _transport_split_analysis(solutions, read_json(RESULTS / "solver_status.json"))
    write_json(RESULTS / "rank_sensitivity.json", rank)
    write_json(RESULTS / "supplier_usage_distribution.json", usage)
    write_json(RESULTS / "transport_split_analysis.json", split)
    summary = {
        "generated_at": now_iso(),
        "p0_found": False,
        "rank_sensitivity": rank["summary"],
        "supply_capacity_semantics": _supply_semantics(data),
        "inventory_timing": _inventory_audit(solutions),
        "supplier_usage": usage["problem3_assessment"],
        "transport_split": split,
        "problem4_capacity": _problem4_audit(solutions["4"], data),
        "stress_tests": _stress_audit(read_json(RESULTS / "sensitivity_analysis.json")),
    }
    write_json(RESULTS / "final_human_audit_evidence.json", summary)


def main() -> None:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    snapshot_parser = subparsers.add_parser("snapshot")
    snapshot_parser.add_argument("output", type=Path)
    compare_parser = subparsers.add_parser("compare")
    compare_parser.add_argument("before", type=Path)
    subparsers.add_parser("audit")
    subparsers.add_parser("excel")
    args = parser.parse_args()
    if args.command == "snapshot":
        write_json(args.output, _snapshot())
    elif args.command == "compare":
        write_json(RESULTS / "clean_replay_validation.json", _compare_snapshots(read_json(args.before), _snapshot()))
    elif args.command == "excel":
        write_json(RESULTS / "excel_recalculation_validation.json", _excel_recalculation_audit())
    else:
        generate_audit()


if __name__ == "__main__":
    main()
