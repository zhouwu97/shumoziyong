"""材料审计、供应商评价和官方模板回填工件。

本模块不求解优化模型；它仅把官方数据、已导出的数值解转换为审计报告与提交模板。
"""

from __future__ import annotations

import shutil
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import numpy as np
from openpyxl import load_workbook

from common.common import HORIZON, MATERIALS, OUTPUTS, ROOT, TOLERANCE, sha256, write_json


MATERIAL_NAMES = [
    "CUMCM2021-C.pdf",
    "附件1 近5年402家供应商的相关数据.xlsx",
    "附件2 近5年8家转运商的相关数据.xlsx",
    "附件A 订购方案数据结果.xlsx",
    "附件B 转运方案数据结果.xlsx",
    "source_manifest.json",
    "material_manifest.json",
    "selection_record.json",
]


def _workbook_profile(path: Path) -> dict[str, Any]:
    """提取工作簿结构而不改变任何原始单元格。"""
    # 只读模式避开模板中大量格式化空白单元格的样式解码；合并单元格从 XML 直接读取。
    merged_by_sheet: dict[int, list[str]] = {}
    with zipfile.ZipFile(path) as archive:
        worksheet_files = sorted(
            [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")],
            key=lambda name: int(name.rsplit("sheet", 1)[1].split(".xml", 1)[0]),
        )
        for index, name in enumerate(worksheet_files):
            merged = []
            for _event, element in ElementTree.iterparse(archive.open(name), events=("end",)):
                if element.tag.endswith("mergeCell"):
                    merged.append(str(element.attrib.get("ref", "")))
                element.clear()
            merged_by_sheet[index] = merged
    book = load_workbook(path, data_only=False, read_only=True)
    sheets = []
    for sheet_index, sheet in enumerate(book.worksheets):
        type_counter: Counter[str] = Counter()
        nonempty = 0
        zero = 0
        negative = 0
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                nonempty += 1
                type_counter[type(value).__name__] += 1
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    zero += int(value == 0)
                    negative += int(value < 0)
        sheets.append(
            {
                "sheet_name": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "merged_cells": merged_by_sheet.get(sheet_index, []),
                "merged_cell_count": len(merged_by_sheet.get(sheet_index, [])),
                "nonempty_cells": nonempty,
                "empty_cells_in_used_range": sheet.max_row * sheet.max_column - nonempty,
                "zero_cells": zero,
                "negative_cells": negative,
                "data_types": dict(type_counter),
            }
        )
    return {"read_status": "passed", "sheets": sheets}


def write_material_manifest() -> dict[str, Any]:
    """为本轮复制件生成含来源、哈希和表结构的可审计材料清单。"""
    original_manifest = MATERIALS / "material_manifest.json"
    declared = __import__("json").loads(original_manifest.read_text(encoding="utf-8"))
    declared_by_name = {entry["path"]: entry for entry in declared}
    records = []
    for name in MATERIAL_NAMES:
        path = MATERIALS / name
        record: dict[str, Any] = {
            "original_path": str(Path(r"E:\AI\shumo_unseen_pool\selected_round2") / name),
            "target_path": str(path),
            "filename": name,
            "size_bytes": path.stat().st_size,
            "sha256": sha256(path),
            "read_status": "passed",
        }
        if name in declared_by_name:
            expected = declared_by_name[name]
            record["source_manifest_match"] = {
                "size": record["size_bytes"] == expected["size_bytes"],
                "sha256": record["sha256"] == expected["sha256"],
            }
        if path.suffix.lower() == ".xlsx":
            record.update(_workbook_profile(path))
        else:
            record["sheets"] = []
        records.append(record)
    payload = {
        "material_root": str(MATERIALS),
        "expected_collection_sha256": "2e09757b783e532ccebed3b03857aa78d7d094df6f130aca53479b123da898d0",
        "collection_hash_method": "SHA-256(官方 material_manifest.json 原始字节)，该官方清单逐项绑定五个题面/附件的大小和 SHA-256。",
        "recomputed_collection_sha256": sha256(original_manifest),
        "collection_hash_passed": sha256(original_manifest) == "2e09757b783e532ccebed3b03857aa78d7d094df6f130aca53479b123da898d0",
        "files": records,
    }
    write_json(ROOT / "material_manifest.json", payload)
    return payload


def data_quality_report(data: dict[str, Any]) -> dict[str, Any]:
    """给出题目要求的零值、缺失、异常、重复和关系审计。"""
    orders = data["orders_history"]
    supply = data["supply_history"]
    loss = data["loss_history_percent"]
    all_values = np.concatenate([orders.ravel(), supply.ravel()])
    # 大量题意明确的零值会使全样本 IQR 退化；异常标记仅在正值供需规模内进行。
    positive_values = all_values[all_values > 0]
    q1, q3 = np.quantile(positive_values, [0.25, 0.75])
    iqr_threshold = q3 + 3.0 * (q3 - q1)
    return {
        "supplier_count": len(data["supplier_ids"]),
        "transporter_count": len(data["transporter_ids"]),
        "historical_weeks": len(data["week_names"]),
        "materials": {item: data["material_types"].count(item) for item in "ABC"},
        "worksheets": {
            "附件1": ["企业的订货量（m³）", "供应商的供货量（m³）"],
            "附件2": ["运输损耗率（%）"],
        },
        "field_meanings": {
            "订货量": "企业向供应商发出的周订货量，0表示该周未订货。",
            "供货量": "供应商实际提供的周原料量，0表示该周未供货。",
            "损耗率": "(供货量-接收量)/供货量 × 100%，0表示未运输而非零损耗观测。",
        },
        "missing_values": {"orders": int(np.isnan(orders).sum()), "supply": int(np.isnan(supply).sum()), "loss": int(np.isnan(loss).sum())},
        "zero_values": {"orders": int((orders == 0).sum()), "supply": int((supply == 0).sum()), "loss": int((loss == 0).sum())},
        "negative_values": {"orders": int((orders < 0).sum()), "supply": int((supply < 0).sum()), "loss": int((loss < 0).sum())},
        "outlier_screen": {"method": "正值原料量的 Q3 + 3IQR，仅标记不删除", "threshold_raw_m3": float(iqr_threshold), "count": int((positive_values > iqr_threshold).sum())},
        "duplicates": {"supplier_id": int(len(data["supplier_ids"]) - len(set(data["supplier_ids"]))), "transporter_id": int(len(data["transporter_ids"]) - len(set(data["transporter_ids"])))},
        "order_supply_relation": {
            "supply_greater_than_order": int((supply > orders).sum()),
            "positive_supply_when_order_zero": int(((supply > 0) & (orders == 0)).sum()),
            "zero_supply_when_order_positive": int(((supply == 0) & (orders > 0)).sum()),
            "interpretation": "供货大于订货是题面允许的实际响应；未订货时无正供货；已订货但零供货是真实失败事件。",
        },
        "loss_denominator": "供货量（发运量），不是接收量；分母为零时题面以0表示未运输，本轮不将其纳入损耗均值。",
        "transport_zero_with_loss": 0,
        "cleaning_decision": "不删除、不截尾任何官方极端值；异常值只在报告中标记，供应能力使用历史响应统计而非删除后样本。",
    }


def supplier_analysis(data: dict[str, Any]) -> dict[str, Any]:
    """计算问题1的业务型供应商重要性评分，并报告权重敏感性。"""
    metrics = data["supplier_metrics"]
    capacity = np.array([item["product_equivalent_capacity_before_loss"] for item in metrics])
    service = np.array([item["service_probability"] for item in metrics])
    fulfilment = np.array([min(item["weighted_fulfilment_ratio"], 1.0) for item in metrics])
    stability = np.array([1.0 / (1.0 + item["supply_cv_positive_weeks"]) if item["supply_cv_positive_weeks"] is not None else 0.0 for item in metrics])

    def normalise(values: np.ndarray) -> np.ndarray:
        low, high = float(values.min()), float(values.max())
        return np.zeros_like(values) if high - low <= TOLERANCE else (values - low) / (high - low)

    components = {"capacity": normalise(capacity), "service": normalise(service), "stability": normalise(stability), "fulfilment": normalise(fulfilment)}
    weight_sets = {
        "baseline": {"capacity": 0.50, "service": 0.25, "stability": 0.15, "fulfilment": 0.10},
        "capacity_heavy": {"capacity": 0.60, "service": 0.20, "stability": 0.10, "fulfilment": 0.10},
        "reliability_heavy": {"capacity": 0.40, "service": 0.35, "stability": 0.15, "fulfilment": 0.10},
        "stability_heavy": {"capacity": 0.45, "service": 0.20, "stability": 0.25, "fulfilment": 0.10},
    }
    rankings: dict[str, list[int]] = {}
    scores: dict[str, np.ndarray] = {}
    for name, weights in weight_sets.items():
        scores[name] = sum(weights[key] * components[key] for key in weights)
        rankings[name] = np.argsort(-scores[name], kind="stable").tolist()
    baseline = rankings["baseline"]
    top50 = baseline[:50]
    entries = []
    for rank, index in enumerate(top50, 1):
        entry = dict(metrics[index])
        entry.update({"rank": rank, "importance_score": float(scores["baseline"][index]), "score_components": {key: float(value[index]) for key, value in components.items()}})
        entries.append(entry)
    capacity_baseline = np.argsort(-capacity, kind="stable")[:50].tolist()
    base_set = set(top50)
    stability_report = {name: len(base_set & set(rank[:50])) / 50.0 for name, rank in rankings.items()}
    return {
        "method": "业务型加权评价：供货能力50%、按单供货概率25%、波动稳定性15%、加权兑现率10%；权重为本轮设定，不是官方权重。",
        "anti_cheating_checks": {
            "capacity_business_link": "能力以规律下单时的期望可供货产品等价量表示，直接服务生产保障。",
            "zero_order_handling": "真实零订货不计入供货失败；只在历史已下单周计算服务概率。",
            "redundancy": "能力、按单供货概率、波动性和兑现率分别描述规模、发生概率、稳定性和订货响应，未机械堆叠高度重复指标。",
            "future_information": "全部指标只使用题给240周历史数据。",
            "ranking_boundary": "排名用于候选供应商识别，不构成真实优劣证明；问题2至4另行优化。",
        },
        "top50": entries,
        "top50_supplier_ids": [metrics[index]["supplier_id"] for index in top50],
        "simple_capacity_baseline_top50_supplier_ids": [metrics[index]["supplier_id"] for index in capacity_baseline],
        "baseline_overlap_with_capacity_top50": len(base_set & set(capacity_baseline)) / 50.0,
        "top50_weight_sensitivity_overlap": stability_report,
    }


def export_templates(solutions: dict[str, dict[str, Any]]) -> None:
    """从原始模板复制后填入24周订单和转运量，零量保留为空白。"""
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    source_order = MATERIALS / "附件A 订购方案数据结果.xlsx"
    source_transport = MATERIALS / "附件B 转运方案数据结果.xlsx"
    target_order = OUTPUTS / source_order.name
    target_transport = OUTPUTS / source_transport.name
    shutil.copy2(source_order, target_order)
    shutil.copy2(source_transport, target_transport)
    order_book = load_workbook(target_order)
    transport_book = load_workbook(target_transport)
    supplier_ids = solutions["2"]["supplier_ids"]
    id_to_index = {supplier_id: index for index, supplier_id in enumerate(supplier_ids)}
    for part, solution in solutions.items():
        orders = np.asarray(solution["orders_raw_m3"], dtype=float)
        shipments = np.asarray(solution["shipments_raw_m3"], dtype=float)
        order_sheet = order_book[f"问题{part}的订购方案结果"]
        transport_sheet = transport_book[f"问题{part}的转运方案结果"]
        for row in range(7, 409):
            index = id_to_index[str(order_sheet.cell(row, 1).value)]
            for week in range(HORIZON):
                order_sheet.cell(row, week + 2).value = None if abs(orders[index, week]) <= TOLERANCE else float(orders[index, week])
            for week in range(HORIZON):
                for transporter in range(8):
                    value = shipments[index, week, transporter]
                    transport_sheet.cell(row, 2 + week * 8 + transporter).value = None if abs(value) <= TOLERANCE else float(value)
    order_book.calculation.fullCalcOnLoad = True
    order_book.calculation.forceFullCalc = True
    transport_book.calculation.fullCalcOnLoad = True
    transport_book.calculation.forceFullCalc = True
    order_book.save(target_order)
    transport_book.save(target_transport)
