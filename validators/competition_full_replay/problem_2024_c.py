"""2024-C 独立复算器；禁止导入题目 Solver 或其目标函数。"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook

from domains.problem_2024_c.data_model import DRYLAND_TYPES, LEGUME_CROP_IDS, YEARS, ProblemData


TOLERANCE = 1e-5


@dataclass(frozen=True)
class WorkbookAssignment:
    plot_id: str
    year: int
    season: str
    crop_id: int
    area_mu: float
    cell: str


def read_official_workbook(path: Path, data: ProblemData) -> list[WorkbookAssignment]:
    """独立按官方行列读取，不调用 Solver 侧导出/导入函数。"""
    book = load_workbook(path, data_only=True, read_only=False)
    expected_sheets = [str(year) for year in YEARS]
    if book.sheetnames != expected_sheets:
        raise ValueError(f"工作表年份不匹配：{book.sheetnames}")
    result: list[WorkbookAssignment] = []
    crop_ids = tuple(sorted(data.crops))
    for year in YEARS:
        sheet = book[str(year)]
        for row in range(2, 84):
            plot_value = sheet.cell(row, 2).value
            if plot_value is None:
                continue
            plot_id = str(plot_value).strip()
            if plot_id not in data.plots:
                raise ValueError(f"{sheet.title}!B{row} 包含未知地块：{plot_id}")
            for offset, crop_id in enumerate(crop_ids, start=3):
                value = sheet.cell(row, offset).value
                if value in (None, ""):
                    continue
                if not isinstance(value, (int, float)):
                    raise ValueError(f"{sheet.title}!{sheet.cell(row, offset).coordinate} 非数值")
                area = float(value)
                if area <= TOLERANCE:
                    continue
                second = row >= 56
                plot_type = data.plots[plot_id].plot_type
                if second:
                    season = "第二季"
                elif plot_type in DRYLAND_TYPES or (plot_type == "水浇地" and crop_id == 16):
                    season = "单季"
                else:
                    season = "第一季"
                result.append(
                    WorkbookAssignment(
                        plot_id,
                        year,
                        season,
                        crop_id,
                        area,
                        f"{sheet.title}!{sheet.cell(row, offset).coordinate}",
                    )
                )
    return result


def _parameter(
    data: ProblemData,
    item: WorkbookAssignment,
    parameter_overrides: dict[tuple[int, int, str, str], dict[str, float]] | None,
) -> dict[str, float]:
    plot_type = data.plots[item.plot_id].plot_type
    key = (item.year, item.crop_id, plot_type, item.season)
    if parameter_overrides is not None:
        return parameter_overrides[key]
    stat = data.stat(item.crop_id, plot_type, item.season)
    return {
        "demand": data.expected_sales_2023.get((item.crop_id, item.season), 0.0),
        "yield": stat.yield_jin_per_mu,
        "cost": stat.cost_yuan_per_mu,
        "price": stat.price_mid_yuan_per_jin,
    }
def recompute_objective(
    assignments: list[WorkbookAssignment],
    data: ProblemData,
    surplus_price_fraction: float,
    parameter_overrides: dict[tuple[int, int, str, str], dict[str, float]] | None = None,
) -> dict[str, object]:
    """从面积、亩产和销售上限重新聚合收入与成本。"""
    production: dict[tuple[int, int, str], float] = defaultdict(float)
    group_parameter: dict[tuple[int, int, str], dict[str, float]] = {}
    cost = 0.0
    for item in assignments:
        parameter = _parameter(data, item, parameter_overrides)
        key = (item.year, item.crop_id, item.season)
        production[key] += item.area_mu * parameter["yield"]
        cost += item.area_mu * parameter["cost"]
        if key in group_parameter:
            previous = group_parameter[key]
            if abs(previous["demand"] - parameter["demand"]) > TOLERANCE or abs(
                previous["price"] - parameter["price"]
            ) > TOLERANCE:
                raise ValueError(f"销售组参数不一致：{key}")
        group_parameter[key] = parameter
    revenue = 0.0
    groups = []
    for key, amount in sorted(production.items()):
        parameter = group_parameter[key]
        normal = min(amount, parameter["demand"])
        surplus = max(0.0, amount - parameter["demand"])
        group_revenue = parameter["price"] * (normal + surplus_price_fraction * surplus)
        revenue += group_revenue
        groups.append(
            {
                "year": key[0],
                "crop_id": key[1],
                "season": key[2],
                "production_jin": amount,
                "normal_sales_jin": normal,
                "surplus_jin": surplus,
                "revenue_yuan": group_revenue,
            }
        )
    return {
        "revenue_yuan": revenue,
        "cost_yuan": cost,
        "profit_yuan": revenue - cost,
        "sales_groups": groups,
    }


def check_constraints(
    assignments: list[WorkbookAssignment],
    data: ProblemData,
    minimum_area_ratio: float = 0.10,
    minimum_area_mu: float = 0.30,
) -> dict[str, object]:
    violations: dict[str, list[dict[str, object]]] = {
        "nonnegative": [],
        "land_capacity": [],
        "crop_eligibility": [],
        "water_mode": [],
        "continuous_cropping": [],
        "three_year_legume": [],
        "minimum_area": [],
    }
    positive: dict[tuple[str, int, str, int], float] = defaultdict(float)
    for item in assignments:
        if item.area_mu < -TOLERANCE:
            violations["nonnegative"].append(asdict(item))
        positive[(item.plot_id, item.year, item.season, item.crop_id)] += item.area_mu
        plot = data.plots[item.plot_id]
        if item.crop_id not in data.eligible_crops(plot.plot_type, item.season):
            violations["crop_eligibility"].append(asdict(item))
        minimum = min(plot.area_mu, max(minimum_area_mu, minimum_area_ratio * plot.area_mu))
        if TOLERANCE < item.area_mu < minimum - TOLERANCE:
            violations["minimum_area"].append(
                {"cell": item.cell, "area_mu": item.area_mu, "minimum_area_mu": minimum}
            )

    used: dict[tuple[str, int, str], float] = defaultdict(float)
    for (plot_id, year, season, _), area in positive.items():
        used[(plot_id, year, season)] += area
    for key, area in used.items():
        capacity = data.plots[key[0]].area_mu
        if area > capacity + TOLERANCE:
            violations["land_capacity"].append(
                {"plot_id": key[0], "year": key[1], "season": key[2], "excess_mu": area - capacity}
            )

    for plot in data.plots.values():
        if plot.plot_type != "水浇地":
            continue
        for year in YEARS:
            rice = used.get((plot.plot_id, year, "单季"), 0.0)
            vegetables = used.get((plot.plot_id, year, "第一季"), 0.0) + used.get(
                (plot.plot_id, year, "第二季"), 0.0
            )
            if rice > TOLERANCE and vegetables > TOLERANCE:
                violations["water_mode"].append({"plot_id": plot.plot_id, "year": year})

    def active(plot_id: str, year: int, season: str) -> set[int]:
        return {
            crop_id
            for (candidate, candidate_year, candidate_season, crop_id), area in positive.items()
            if candidate == plot_id
            and candidate_year == year
            and candidate_season == season
            and area > TOLERANCE
        }

    for plot in data.plots.values():
        pairs: list[tuple[int, str, int, str]] = []
        if plot.plot_type in DRYLAND_TYPES:
            pairs.extend((year, "单季", year + 1, "单季") for year in YEARS[:-1])
            boundary = (("单季", "单季"),)
        elif plot.plot_type == "水浇地":
            pairs.extend((year, "单季", year + 1, "单季") for year in YEARS[:-1])
            boundary = (("单季", "单季"),)
        elif plot.plot_type == "智慧大棚":
            for year in YEARS:
                pairs.append((year, "第一季", year, "第二季"))
                if year < YEARS[-1]:
                    pairs.append((year, "第二季", year + 1, "第一季"))
            boundary = (("第二季", "第一季"),)
        else:
            boundary = ()
        for left_year, left_season, right_year, right_season in pairs:
            for crop_id in active(plot.plot_id, left_year, left_season) & active(
                plot.plot_id, right_year, right_season
            ):
                violations["continuous_cropping"].append(
                    {
                        "plot_id": plot.plot_id,
                        "crop_id": crop_id,
                        "left": f"{left_year}-{left_season}",
                        "right": f"{right_year}-{right_season}",
                    }
                )
        for historical_season, planned_season in boundary:
            historical = {item.crop_id for item in data.history_for(plot.plot_id, historical_season)}
            for crop_id in historical & active(plot.plot_id, 2024, planned_season):
                violations["continuous_cropping"].append(
                    {
                        "plot_id": plot.plot_id,
                        "crop_id": crop_id,
                        "left": f"2023-{historical_season}",
                        "right": f"2024-{planned_season}",
                    }
                )

        history_legume = sum(
            item.area_mu for item in data.history_for(plot.plot_id) if item.crop_id in LEGUME_CROP_IDS
        )
        for start in range(2023, 2029):
            amount = history_legume if start == 2023 else 0.0
            for (plot_id, year, _season, crop_id), area in positive.items():
                if plot_id == plot.plot_id and start <= year <= start + 2 and crop_id in LEGUME_CROP_IDS:
                    amount += area
            if amount < plot.area_mu - TOLERANCE:
                violations["three_year_legume"].append(
                    {
                        "plot_id": plot.plot_id,
                        "window": f"{start}-{start + 2}",
                        "shortfall_mu": plot.area_mu - amount,
                    }
                )

    return {
        "passed": all(not items for items in violations.values()),
        "violation_counts": {name: len(items) for name, items in violations.items()},
        "violations": violations,
        "assignment_count": len(assignments),
        "total_area_mu": sum(item.area_mu for item in assignments),
    }


def validate_q1_workbook(
    workbook: Path,
    data: ProblemData,
    scenario_id: str,
    objective_reported: float,
) -> dict[str, object]:
    if scenario_id not in {"q1_waste", "q1_discount"}:
        raise ValueError(f"未知 Q1 场景：{scenario_id}")
    assignments = read_official_workbook(workbook, data)
    fraction = 0.0 if scenario_id == "q1_waste" else 0.5
    objective = recompute_objective(assignments, data, fraction)
    constraints = check_constraints(assignments, data)
    objective_data = cast(dict[str, Any], objective)
    error = abs(float(objective_data["profit_yuan"]) - objective_reported)
    return {
        "schema_version": "1.0.0",
        "artifact_type": "2024_c_problem_specific_validation",
        "problem_id": "2024-C",
        "scenario_id": scenario_id,
        "workbook": str(workbook),
        "objective_reported_yuan": objective_reported,
        "objective_recomputed_yuan": objective["profit_yuan"],
        "objective_absolute_error_yuan": error,
        "objective_tolerance_yuan": 1e-3,
        "objective_passed": error <= 1e-3,
        "constraints": constraints,
        "passed": error <= 1e-3 and bool(constraints["passed"]),
    }
