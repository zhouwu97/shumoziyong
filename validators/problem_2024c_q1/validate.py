"""2024-C Q1 的官方数据加载、目标复算和硬约束检查。

本模块只实现 Q1 的两种销售口径，不包含 Q2/Q3 的不确定性或相关性语义。
"""

from __future__ import annotations

import hashlib
import json
import math
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping

from jsonschema import Draft202012Validator
from openpyxl import load_workbook


YEARS = tuple(range(2024, 2031))
Q1_SCENARIOS = ("q1_waste", "q1_discount")
LEGUME_CROPS = frozenset({1, 2, 3, 4, 5, 17, 18, 19})
TOLERANCE = 1e-6
FORMAL_MATERIAL_MANIFEST_SCHEMA_PATH = (
    Path(__file__).parents[2] / "schemas" / "2024c_official_material_manifest.schema.json"
)
ATTACHMENT_ROLES = {
    "attachment_1": ("land_and_crop_dictionary", "attachments/附件1.xlsx"),
    "attachment_2": ("historical_planting_and_statistics", "attachments/附件2.xlsx"),
}


def _as_float(value: object) -> float:
    if isinstance(value, (int, float, str)) and not isinstance(value, bool):
        result = float(value)
        if math.isfinite(result):
            return result
    raise ValueError(f"无法转换为有限数值: {value!r}")


def _price_midpoint(value: object) -> float:
    text = str(value).strip()
    if "-" in text:
        left, right = text.split("-", 1)
        return (_as_float(left) + _as_float(right)) / 2.0
    return _as_float(text)


def load_q1_data(attachment_1: Path, attachment_2: Path) -> dict[str, Any]:
    """从官方附件恢复合并单元格并冻结 Q1 所需数据。"""

    book_1 = load_workbook(attachment_1, read_only=True, data_only=True)
    land_sheet = book_1["乡村的现有耕地"]
    plots = {
        str(row[0]).strip(): {"type": str(row[1]).strip(), "area": _as_float(row[2])}
        for row in land_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }

    book_2 = load_workbook(attachment_2, read_only=True, data_only=True)
    stat_sheet = book_2["2023年统计的相关数据"]
    stats: dict[tuple[str, str, int], dict[str, float]] = {}
    for row in stat_sheet.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[1], (int, float)):
            continue
        key = (str(row[3]).strip(), str(row[4]).strip(), int(row[1]))
        stats[key] = {"yield": _as_float(row[5]), "cost": _as_float(row[6]), "price": _price_midpoint(row[7])}
    for crop_id in range(17, 35):
        source = stats.get(("普通大棚", "第一季", crop_id))
        if source is None:
            raise ValueError(f"缺少普通大棚第一季作物参数: {crop_id}")
        stats[("智慧大棚", "第一季", crop_id)] = dict(source)

    planting: list[dict[str, Any]] = []
    current_plot: str | None = None
    production: defaultdict[tuple[int, str], float] = defaultdict(float)
    for row in book_2["2023年的农作物种植情况"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            current_plot = str(row[0]).strip()
        if current_plot is None or not isinstance(row[1], (int, float)):
            continue
        item = {
            "year": 2023,
            "plot_id": current_plot,
            "season": str(row[5]).strip(),
            "crop_id": int(row[1]),
            "area_mu": _as_float(row[4]),
        }
        if current_plot not in plots:
            raise ValueError(f"2023 种植记录引用未知地块: {current_plot}")
        stat = stats.get((plots[current_plot]["type"], item["season"], item["crop_id"]))
        if stat is None:
            raise ValueError(f"2023 记录缺少统计参数: {item}")
        production[(item["crop_id"], item["season"])] += item["area_mu"] * stat["yield"]
        planting.append(item)

    prices: defaultdict[tuple[int, str], set[float]] = defaultdict(set)
    for (plot_type, season, crop_id), stat in stats.items():
        prices[(crop_id, season)].add(stat["price"])
    inconsistent = {key: sorted(values) for key, values in prices.items() if len(values) != 1}
    if inconsistent:
        raise ValueError(f"同一作物-季次销售价格不唯一: {inconsistent}")
    return {
        "plots": plots,
        "stats": stats,
        "planting_2023": planting,
        "sales_2023": dict(production),
        "price_by_crop_season": {key: next(iter(values)) for key, values in prices.items()},
    }


def _items(assignments: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "year": int(item["year"]),
            "plot_id": str(item["plot_id"]).strip(),
            "season": str(item["season"]).strip(),
            "crop_id": int(item["crop_id"]),
            "area_mu": float(item["area_mu"]),
        }
        for item in assignments
    ]


def evaluate_q1_objective(assignments: Iterable[Mapping[str, Any]], data: Mapping[str, Any], scenario: str) -> float:
    """按 Q1 销售上限和成本/价格口径独立复算利润。"""

    if scenario not in Q1_SCENARIOS:
        raise ValueError(f"Q1 不接受场景: {scenario}")
    production: defaultdict[tuple[int, str, int], float] = defaultdict(float)
    costs = 0.0
    for item in _items(assignments):
        if item["plot_id"] not in data["plots"]:
            raise ValueError(f"未知地块: {item['plot_id']}")
        plot = data["plots"][item["plot_id"]]
        stat = data["stats"][(plot["type"], item["season"], item["crop_id"])]
        production[(item["crop_id"], item["season"], item["year"])] += item["area_mu"] * stat["yield"]
        costs += item["area_mu"] * stat["cost"]

    discount = 0.5 if scenario == "q1_discount" else 0.0
    revenue = 0.0
    for (crop_id, season, year), amount in production.items():
        cap = data["sales_2023"].get((crop_id, season), 0.0)
        price = data["price_by_crop_season"][(crop_id, season)]
        sold = min(amount, cap)
        revenue += sold * price + max(amount - cap, 0.0) * price * discount
    return revenue - costs


def _continuous_violations(items: list[dict[str, Any]], data: Mapping[str, Any]) -> list[str]:
    presence = {(item["plot_id"], item["year"], item["season"], item["crop_id"]) for item in items if item["area_mu"] > TOLERANCE}
    presence.update((item["plot_id"], 2023, item["season"], item["crop_id"]) for item in data["planting_2023"] if item["area_mu"] > TOLERANCE)
    violations: list[str] = []
    for plot_id, plot in data["plots"].items():
        if plot["type"] == "智慧大棚":
            slots = [(year, season) for year in range(2023, 2031) for season in ("第一季", "第二季")]
            pairs = zip(slots, slots[1:])
        elif plot["type"] in {"平旱地", "梯田", "山坡地", "水浇地"}:
            pairs = (
                ((year - 1, "单季"), (year, "单季"))
                for year in range(2024, 2031)
            )
        else:
            # 普通大棚相邻季为蔬菜与食用菌，适种集合不相交，不会同作物重茬。
            pairs = iter(())
        for previous, current in pairs:
            for crop_id in range(1, 42):
                if (plot_id, previous[0], previous[1], crop_id) in presence and (plot_id, current[0], current[1], crop_id) in presence:
                    violations.append(f"continuous_crop:{plot_id}:{crop_id}:{previous[0]}-{previous[1]}->{current[0]}-{current[1]}")
    return violations


def check_q1_constraints(assignments: Iterable[Mapping[str, Any]], data: Mapping[str, Any], *, check_legume_windows: bool = True) -> tuple[list[str], float]:
    """检查 Q1 年份、适宜性、容量、种植制度、重茬和豆类窗口。"""

    items = _items(assignments)
    violations: list[str] = []
    max_violation = 0.0
    grouped: defaultdict[tuple[int, str, str], float] = defaultdict(float)
    crop_year_area: defaultdict[tuple[str, int, int], float] = defaultdict(float)
    valid: list[dict[str, Any]] = []
    for item in items:
        if item["year"] not in YEARS:
            violations.append(f"year:{item['year']}")
            continue
        if item["plot_id"] not in data["plots"]:
            violations.append(f"plot:{item['plot_id']}")
            continue
        if not math.isfinite(item["area_mu"]) or item["area_mu"] < -TOLERANCE:
            violations.append(f"negative_or_nonfinite_area:{item['plot_id']}")
        plot = data["plots"][item["plot_id"]]
        if (plot["type"], item["season"], item["crop_id"]) not in data["stats"]:
            violations.append(f"suitability:{item['plot_id']}:{item['season']}:{item['crop_id']}")
            continue
        valid.append(item)
        grouped[(item["year"], item["plot_id"], item["season"])] += item["area_mu"]
        crop_year_area[(item["plot_id"], item["crop_id"], item["year"])] += item["area_mu"]

    for (year, plot_id, season), area in grouped.items():
        violation = max(area - float(data["plots"][plot_id]["area"]), 0.0)
        if violation > TOLERANCE:
            violations.append(f"capacity:{year}:{plot_id}:{season}")
        max_violation = max(max_violation, violation)

    for plot_id, plot in data["plots"].items():
        if plot["type"] == "水浇地":
            for year in YEARS:
                single = grouped[(year, plot_id, "单季")]
                for season in ("第一季", "第二季"):
                    violation = max(single + grouped[(year, plot_id, season)] - float(plot["area"]), 0.0)
                    if violation > TOLERANCE:
                        violations.append(f"water_system:{year}:{plot_id}:{season}")
                    max_violation = max(max_violation, violation)

    violations.extend(_continuous_violations(valid, data))
    if check_legume_windows:
        for item in data["planting_2023"]:
            crop_year_area[(item["plot_id"], item["crop_id"], 2023)] += item["area_mu"]
        for plot_id, plot in data["plots"].items():
            for start in range(2023, 2029):
                legume_area = sum(crop_year_area[(plot_id, crop_id, year)] for crop_id in LEGUME_CROPS for year in range(start, start + 3))
                violation = max(float(plot["area"]) - legume_area, 0.0)
                if violation > TOLERANCE:
                    violations.append(f"legume_window:{plot_id}:{start}-{start + 2}")
                max_violation = max(max_violation, violation)
    return sorted(set(violations)), max_violation


def _load_bound_material_manifest(
    material_manifest: Path,
    attachment_1: Path,
    attachment_2: Path,
) -> dict[str, Any]:
    """验证材料清单身份，并将两个输入附件绑定到清单声明的文件。"""

    try:
        manifest = json.loads(material_manifest.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"材料 Manifest 无法读取: {material_manifest}") from exc
    schema = json.loads(FORMAL_MATERIAL_MANIFEST_SCHEMA_PATH.read_text(encoding="utf-8"))
    schema_errors = list(Draft202012Validator(schema).iter_errors(manifest))
    if schema_errors:
        raise ValueError("材料 Manifest Schema 无效: " + "; ".join(error.message for error in schema_errors))
    if manifest["problem_id"] != "2024-C":
        raise ValueError(f"材料 Manifest 题目不匹配: {manifest['problem_id']!r}")
    if manifest["source"]["kind"] != "official":
        raise ValueError("Q1 Validator 只接受官方材料 Manifest")

    records: dict[str, dict[str, Any]] = {}
    paths: set[str] = set()
    for record in manifest["files"]:
        role = record["role"]
        path = record["path"]
        if role in records or path in paths:
            raise ValueError(f"正式材料 Manifest 存在重复角色或路径: {role} ({path})")
        records[role] = record
        paths.add(path)

    attachments = {"attachment_1": attachment_1, "attachment_2": attachment_2}
    for attachment_role, actual_path in attachments.items():
        manifest_role, relative_path = ATTACHMENT_ROLES[attachment_role]
        record = records.get(manifest_role)
        if record is None:
            raise ValueError(f"正式材料 Manifest 缺少附件角色: {manifest_role}")
        if record["path"] != relative_path or actual_path.name != Path(relative_path).name:
            raise ValueError(f"实际附件与正式 Manifest 角色不匹配: {attachment_role}")
        if not actual_path.is_file():
            raise ValueError(f"Manifest 声明的附件不存在: {actual_path}")
        if actual_path.stat().st_size != record["bytes"]:
            raise ValueError(f"附件字节数与正式 Manifest 不匹配: {relative_path}")
        actual_sha = hashlib.sha256(actual_path.read_bytes()).hexdigest()
        if actual_sha != record["sha256"]:
            raise ValueError(f"附件 SHA-256 与正式 Manifest 不匹配: {relative_path}")
    return manifest


def validate_q1_result(result: Mapping[str, Any], attachment_1: Path, attachment_2: Path, material_manifest: Path, *, check_legume_windows: bool = True, objective_tolerance: float = TOLERANCE) -> dict[str, Any]:
    """验证 Q1 两个场景；缺场景、目标漂移或硬约束失败时 fail-closed。"""

    schema = json.loads((Path(__file__).parents[2] / "schemas" / "2024c_q1_result.schema.json").read_text(encoding="utf-8"))
    schema_errors = list(Draft202012Validator(schema).iter_errors(result))
    if schema_errors:
        raise ValueError("Q1 Formal Result Schema 无效: " + "; ".join(error.message for error in schema_errors))
    actual_manifest_sha = hashlib.sha256(material_manifest.read_bytes()).hexdigest()
    if result["material_manifest_sha256"] != actual_manifest_sha:
        raise ValueError("材料 Manifest SHA-256 不匹配")
    _load_bound_material_manifest(material_manifest, attachment_1, attachment_2)
    data = load_q1_data(attachment_1, attachment_2)
    reports: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in result["scenarios"]:
        scenario = item["scenario_id"]
        if scenario in seen:
            raise ValueError(f"Q1 场景重复: {scenario}")
        seen.add(scenario)
        recomputed = evaluate_q1_objective(item["assignments"], data, scenario)
        violations, max_violation = check_q1_constraints(item["assignments"], data, check_legume_windows=check_legume_windows)
        difference = abs(recomputed - float(item["objective_reported"]))
        reports.append({"scenario_id": scenario, "objective_recomputed": recomputed, "objective_reported": item["objective_reported"], "objective_difference": difference, "violated_constraints": violations, "max_raw_constraint_violation": max_violation, "output_workbook_status": item["output_workbook_status"], "valid": math.isfinite(difference) and difference <= objective_tolerance and not violations})
    valid = len(reports) == 2 and all(report["valid"] for report in reports)
    return {
        "validator": "2024-c-q1-independent-validator-v1",
        "problem_id": "2024-C",
        "q1_status": "implemented",
        "reports": reports,
        "valid": valid,
        "production_ready": False,
    }
