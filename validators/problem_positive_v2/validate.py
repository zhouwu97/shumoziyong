"""2024-C 经官方附件预处理复核后的目标与约束适配器。"""

from __future__ import annotations

import math
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook


YEARS = tuple(range(2024, 2031))
SCENARIOS = ("q1_waste", "q1_discount", "q2_frozen", "q3_frozen")
LEGUME_CROPS = frozenset({1, 2, 3, 4, 5, 17, 18, 19})


def _price_midpoint(value: object) -> float:
    text = str(value).strip()
    if "-" in text:
        left, right = text.split("-", 1)
        return (float(left) + float(right)) / 2.0
    return float(text)


def _as_float(value: object) -> float:
    if isinstance(value, (int, float, str)):
        return float(value)
    raise ValueError(f"无法转换为数值: {value!r}")


def load_problem_data(attachment_1: Path, attachment_2: Path) -> dict[str, Any]:
    """读取官方附件，并恢复合并单元格与智慧大棚第一季参数。"""
    book_1 = load_workbook(attachment_1, read_only=True, data_only=True)
    land_sheet = book_1["乡村的现有耕地"]
    plots = {
        str(row[0]).strip(): {"type": str(row[1]).strip(), "area": _as_float(row[2])}
        for row in land_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }

    book_2 = load_workbook(attachment_2, read_only=True, data_only=True)
    stat_sheet = book_2["2023年统计的相关数据"]
    stats: dict[tuple[str, str, int], dict[str, float]] = {}
    for row in stat_sheet.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[1], (int, float)):
            continue
        crop_id = int(row[1])
        plot_type = str(row[3]).strip()
        season = str(row[4]).strip()
        stats[(plot_type, season, crop_id)] = {
            "yield": _as_float(row[5]),
            "cost": _as_float(row[6]),
            "price": _price_midpoint(row[7]),
        }
    for crop_id in range(17, 35):
        stats[("智慧大棚", "第一季", crop_id)] = dict(
            stats[("普通大棚", "第一季", crop_id)]
        )

    planting_sheet = book_2["2023年的农作物种植情况"]
    planting_2023: list[dict[str, Any]] = []
    production_2023: defaultdict[tuple[int, str], float] = defaultdict(float)
    current_plot: str | None = None
    for row in planting_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            current_plot = str(row[0]).strip()
        if current_plot is None or not isinstance(row[1], (int, float)):
            continue
        crop_id = int(row[1])
        area = _as_float(row[4])
        season = str(row[5]).strip()
        planting_2023.append(
            {
                "year": 2023,
                "plot_id": current_plot,
                "season": season,
                "crop_id": crop_id,
                "area_mu": area,
            }
        )
        plot_type = plots[current_plot]["type"]
        stat = stats[(plot_type, season, crop_id)]
        production_2023[(crop_id, season)] += area * stat["yield"]

    prices: defaultdict[tuple[int, str], set[float]] = defaultdict(set)
    for (_plot_type, season, crop_id), stat in stats.items():
        prices[(crop_id, season)].add(stat["price"])
    inconsistent = {
        key: sorted(values) for key, values in prices.items() if len(values) != 1
    }
    if inconsistent:
        raise ValueError(f"同一作物-季次存在不一致销售价格: {inconsistent}")
    price_by_crop_season = {key: next(iter(values)) for key, values in prices.items()}
    return {
        "plots": plots,
        "stats": stats,
        "planting_2023": planting_2023,
        "sales_2023": dict(production_2023),
        "price_by_crop_season": price_by_crop_season,
        "preprocessing_contract": {
            "merged_plot_ids_forward_filled": True,
            "smart_greenhouse_first_season_source": "普通大棚第一季",
            "sales_cap_key": ["crop_id", "season"],
        },
    }


def _factors(scenario: str, crop_id: int, year: int) -> tuple[float, float, float, float]:
    step = year - 2023
    sales = 1.0
    yield_factor = 1.0
    cost = 1.0
    price = 1.0
    if scenario in {"q2_frozen", "q3_frozen"}:
        sales = 1.075**step if crop_id in {6, 7} else 1.0
        yield_factor = 0.95
        cost = 1.05**step
        if 17 <= crop_id <= 37:
            price = 1.05**step
        elif crop_id == 41:
            price = 0.95**step
        elif 38 <= crop_id <= 40:
            price = 0.97**step
    return sales, yield_factor, cost, price


def _normalize_assignments(items: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "year": int(item["year"]),
            "plot_id": str(item["plot_id"]).strip(),
            "season": str(item["season"]).strip(),
            "crop_id": int(item["crop_id"]),
            "area_mu": float(item["area_mu"]),
        }
        for item in items
    ]


def evaluate_objective(
    assignments: Iterable[Mapping[str, Any]], data: Mapping[str, Any], scenario: str
) -> float:
    """按作物-季次销售上限复算七年利润。"""
    production: defaultdict[tuple[int, str, int], float] = defaultdict(float)
    costs = 0.0
    for item in _normalize_assignments(assignments):
        plot = data["plots"][item["plot_id"]]
        stat = data["stats"][(plot["type"], item["season"], item["crop_id"])]
        _, yield_factor, cost_factor, _ = _factors(
            scenario, item["crop_id"], item["year"]
        )
        production[(item["crop_id"], item["season"], item["year"])] += (
            item["area_mu"] * stat["yield"] * yield_factor
        )
        costs += item["area_mu"] * stat["cost"] * cost_factor

    revenue = 0.0
    discount = 0.0 if scenario == "q1_waste" else 0.5 if scenario == "q1_discount" else 0.0
    for (crop_id, season, year), amount in production.items():
        sales_factor, _, _, price_factor = _factors(scenario, crop_id, year)
        cap = data["sales_2023"].get((crop_id, season), 0.0) * sales_factor
        if "price_by_crop_season" in data:
            base_price = data["price_by_crop_season"][(crop_id, season)]
        else:
            candidates = {
                stat["price"]
                for (_plot_type, stat_season, stat_crop), stat in data["stats"].items()
                if stat_season == season and stat_crop == crop_id
            }
            if len(candidates) != 1:
                raise ValueError(f"无法确定唯一价格: crop={crop_id}, season={season}")
            base_price = next(iter(candidates))
        price = base_price * price_factor
        sold = min(amount, cap)
        excess = max(amount - cap, 0.0)
        revenue += sold * price + excess * price * discount
    return revenue - costs


def _continuous_crop_violations(
    items: list[dict[str, Any]], data: Mapping[str, Any], tolerance: float
) -> list[str]:
    presence = {
        (item["plot_id"], item["year"], item["season"], item["crop_id"])
        for item in items
        if item["area_mu"] > tolerance
    }
    presence.update(
        (item["plot_id"], 2023, item["season"], item["crop_id"])
        for item in data.get("planting_2023", [])
        if item["area_mu"] > tolerance
    )
    violations: list[str] = []
    for plot_id, plot in data["plots"].items():
        if plot["type"] == "智慧大棚":
            slots = [
                (year, season)
                for year in range(2023, 2031)
                for season in ("第一季", "第二季")
            ]
            for previous, current in zip(slots, slots[1:]):
                for crop_id in range(17, 35):
                    if (
                        (plot_id, previous[0], previous[1], crop_id) in presence
                        and (plot_id, current[0], current[1], crop_id) in presence
                    ):
                        violations.append(
                            f"continuous_crop:{plot_id}:{crop_id}:"
                            f"{previous[0]}-{previous[1]}->{current[0]}-{current[1]}"
                        )
        else:
            seasons = {season for p, _year, season, _crop in presence if p == plot_id}
            for season in seasons:
                for crop_id in range(1, 42):
                    for year in YEARS:
                        if (
                            (plot_id, year - 1, season, crop_id) in presence
                            and (plot_id, year, season, crop_id) in presence
                        ):
                            violations.append(
                                f"continuous_crop:{plot_id}:{season}:{crop_id}:{year}"
                            )
    return violations


def check_constraints(
    assignments: Iterable[Mapping[str, Any]],
    data: Mapping[str, Any],
    tolerance: float,
    *,
    check_legume_windows: bool = True,
) -> tuple[list[str], float]:
    """检查适宜性、容量、轮作和三年豆类覆盖。"""
    items = _normalize_assignments(assignments)
    violations: list[str] = []
    max_violation = 0.0
    grouped: defaultdict[tuple[int, str, str], float] = defaultdict(float)
    crop_year_area: defaultdict[tuple[str, int, int], float] = defaultdict(float)
    valid_items: list[dict[str, Any]] = []
    for item in items:
        if item["year"] not in YEARS:
            violations.append(f"year:{item['year']}")
            continue
        if item["plot_id"] not in data["plots"]:
            violations.append(f"plot:{item['plot_id']}")
            continue
        if item["area_mu"] < -tolerance:
            violations.append(f"negative_area:{item['plot_id']}")
        plot = data["plots"][item["plot_id"]]
        if (plot["type"], item["season"], item["crop_id"]) not in data["stats"]:
            violations.append(
                f"suitability:{item['plot_id']}:{item['season']}:{item['crop_id']}"
            )
            continue
        valid_items.append(item)
        grouped[(item["year"], item["plot_id"], item["season"])] += item["area_mu"]
        crop_year_area[(item["plot_id"], item["crop_id"], item["year"])] += item[
            "area_mu"
        ]

    for (year, plot_id, season), area in grouped.items():
        capacity = float(data["plots"][plot_id]["area"])
        violation = max(area - capacity, 0.0)
        if violation > tolerance:
            violations.append(f"capacity:{year}:{plot_id}:{season}")
        max_violation = max(max_violation, violation)

    for plot_id, plot in data["plots"].items():
        if plot["type"] == "水浇地":
            for year in YEARS:
                rice = grouped[(year, plot_id, "单季")]
                for season in ("第一季", "第二季"):
                    violation = max(rice + grouped[(year, plot_id, season)] - plot["area"], 0.0)
                    if violation > tolerance:
                        violations.append(f"water_system:{year}:{plot_id}:{season}")
                    max_violation = max(max_violation, violation)

    violations.extend(_continuous_crop_violations(valid_items, data, tolerance))
    if check_legume_windows:
        for item in data.get("planting_2023", []):
            crop_year_area[(item["plot_id"], item["crop_id"], 2023)] += item["area_mu"]
        for plot_id in data["plots"]:
            for start in range(2023, 2029):
                legume_area = sum(
                    crop_year_area[(plot_id, crop_id, year)]
                    for crop_id in LEGUME_CROPS
                    for year in range(start, start + 3)
                )
                violation = max(float(data["plots"][plot_id]["area"]) - legume_area, 0.0)
                if violation > tolerance:
                    violations.append(f"legume_window:{plot_id}:{start}-{start + 2}")
                max_violation = max(max_violation, violation)
    return sorted(set(violations)), max_violation


def validate_result(
    result: Mapping[str, Any],
    attachment_1: Path,
    attachment_2: Path,
    *,
    objective_tolerance: float = 1e-6,
    constraint_tolerance: float = 1e-5,
) -> dict[str, Any]:
    """验证四个冻结场景的目标和共同可行性。"""
    data = load_problem_data(attachment_1, attachment_2)
    scenario_reports: list[dict[str, Any]] = []
    all_valid = True
    by_id = {str(item["scenario_id"]): item for item in result.get("scenarios", [])}
    for scenario in SCENARIOS:
        item = by_id.get(scenario)
        if item is None:
            scenario_reports.append({"scenario_id": scenario, "valid": False, "reason": "missing"})
            all_valid = False
            continue
        assignments = item.get("assignments", [])
        recomputed = evaluate_objective(assignments, data, scenario)
        reported = float(item.get("objective_reported", math.nan))
        difference = abs(recomputed - reported)
        violations, max_violation = check_constraints(assignments, data, constraint_tolerance)
        objective_valid = math.isfinite(difference) and difference <= objective_tolerance
        constraints_valid = not violations
        valid = objective_valid and constraints_valid
        all_valid = all_valid and valid
        scenario_reports.append(
            {
                "scenario_id": scenario,
                "objective_reported": reported,
                "objective_recomputed": recomputed,
                "objective_difference": difference,
                "objective_valid": objective_valid,
                "violated_constraints": violations,
                "max_raw_constraint_violation": max_violation,
                "constraints_valid": constraints_valid,
                "valid": valid,
            }
        )
    return {
        "validator": "a092_2024c_full_v2",
        "data_contract_version": "2024c_official_attachments_v2",
        "objective_tolerance": objective_tolerance,
        "constraint_tolerance": constraint_tolerance,
        "scenario_reports": scenario_reports,
        "q3_correlation_claim_requires_separate_evidence": True,
        "valid": all_valid,
    }
