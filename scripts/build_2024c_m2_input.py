"""构造 2024-C 单季旱地基线的 Collector 白名单输入。"""
from __future__ import annotations
import argparse, hashlib, json, shutil, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CODE = '''from pathlib import Path
import json
from openpyxl import load_workbook
land=load_workbook(Path("../materials/附件1.xlsx"),data_only=True,read_only=True)["乡村的现有耕地"]
stats=load_workbook(Path("../materials/附件2.xlsx"),data_only=True,read_only=True)["2023年统计的相关数据"]
profits={}
for r in stats.iter_rows(min_row=2,values_only=True):
    if isinstance(r[1],int) and 1<=r[1]<=15 and r[3] in {"平旱地","梯田","山坡地"} and r[4]=="单季" and isinstance(r[7],str):
        lo,hi=map(float,r[7].strip().split("-")); profits[(r[1],r[3])]=r[5]*(lo+hi)/2-r[6]
assign=[]; objective=0.0
for r in land.iter_rows(min_row=2,values_only=True):
    if r[1] in {"平旱地","梯田","山坡地"}:
        crop=max((c for c in range(1,16) if (c,r[1]) in profits),key=lambda c:(profits[(c,r[1])],-c)); assign.append({"plot_id":r[0],"crop_id":crop,"area_mu":float(r[2])}); objective+=float(r[2])*profits[(crop,r[1])]
Path("output").mkdir(exist_ok=True); Path("output/decision_variables.json").write_text(json.dumps({"schema_version":"1.0.0","problem_id":"2024-C","scope":"q1_dryland_single_season_baseline","task_id":"Q1_DRYLAND_BASELINE","assignments":assign,"objective_reported":objective},ensure_ascii=False),encoding="utf-8")
'''
def sha(p:Path)->str:return hashlib.sha256(p.read_bytes()).hexdigest()
def main():
 p=argparse.ArgumentParser();p.add_argument('--materials',type=Path,required=True);p.add_argument('--output',type=Path,required=True);a=p.parse_args();o=a.output;o.mkdir(parents=True,exist_ok=False);(o/'workspace/code').mkdir(parents=True);(o/'materials').mkdir();
 for n in ['附件1.xlsx','附件2.xlsx']:shutil.copy2(a.materials/'attachments'/n,o/'materials'/n)
 shutil.copy2(a.materials/'material_manifest.json',o/'materials'/'material_manifest.json')
 (o/'workspace/code/q1_dryland_baseline.py').write_text(CODE,encoding='utf-8'); (o/'model_route_v2.json').write_text('{"scope":"baseline"}',encoding='utf-8');(o/'environment_lock.json').write_text(json.dumps({'python_executable':sys.executable,'solver':{'name':'deterministic_greedy','version':'1.0','status':'optimal'}}),encoding='utf-8')
 spec={'network_access':False,'tasks':[{'task_id':'Q1_DRYLAND_BASELINE','argv':['python','code/q1_dryland_baseline.py'],'timeout_seconds':120}]};(o/'execution_spec.json').write_text(json.dumps(spec),encoding='utf-8')
if __name__=='__main__':main()
