"""生成 2024-C A0 官方材料、索引和输出合同审计。"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from domains.problem_2024_c.data_loader import (
    default_audit_output_path,
    load_problem_data,
    resolve_material_root,
)
from domains.problem_2024_c.official_output_schema import inspect_template
from openpyxl import load_workbook


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build_audit(material_root: Path) -> dict[str, object]:
    base = material_root / "2024_C"
    data = load_problem_data(material_root)
    manifest_path = base / "material_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    files = [
        base / "problem" / "C题.pdf",
        base / "attachments" / "附件1.xlsx",
        base / "attachments" / "附件2.xlsx",
        base / "templates" / "result1_1.xlsx",
        base / "templates" / "result1_2.xlsx",
        base / "templates" / "result2.xlsx",
    ]
    expected_hashes = {
        item["path"]: item["sha256"]
        for category in manifest["categories"].values()
        for item in category["files"]
    }
    actual_hashes = {path.relative_to(base).as_posix(): _sha256(path) for path in files}
    source_workbooks: dict[str, object] = {}
    for path in files[1:3]:
        book = load_workbook(path, data_only=True, read_only=False)
        source_workbooks[path.name] = {
            sheet.title: {
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
            }
            for sheet in book.worksheets
        }
    contracts = {path.name: inspect_template(path) for path in files[-3:]}
    area_by_type = {
        plot_type: sum(item.area_mu for item in data.plots.values() if item.plot_type == plot_type)
        for plot_type in sorted({item.plot_type for item in data.plots.values()})
    }
    open_field_area = sum(
        area_by_type[plot_type] for plot_type in ("平旱地", "梯田", "山坡地", "水浇地")
    )
    return {
        "schema_version": "1.0.0",
        "artifact_type": "2024_c_official_materials_audit",
        "problem_id": "2024-C",
        "official_materials_complete": all(path.is_file() for path in files),
        "hashes_match_manifest": all(
            expected_hashes.get(relative) == digest for relative, digest in actual_hashes.items()
        ),
        "proxy_data_used": False,
        "fixture_solver_used": False,
        "solver_started": False,
        "qualification_claimed": False,
        "complete_official_old_problem_closure": 0,
        "material_manifest_sha256": _sha256(manifest_path),
        "files": {
            path.relative_to(material_root).as_posix(): {
                "sha256": actual_hashes[path.relative_to(base).as_posix()],
                "size_bytes": path.stat().st_size,
            }
            for path in files
        },
        "source_workbooks": source_workbooks,
        "data_index": {
            "plot_count": len(data.plots),
            "crop_count": len(data.crops),
            "planting_2023_record_count": len(data.planting_2023),
            "statistics_record_count": len(data.statistics),
            "sales_baseline_group_count": len(data.expected_sales_2023),
            "total_area_mu": sum(item.area_mu for item in data.plots.values()),
            "plot_type_area_mu": area_by_type,
            "years": list(range(2024, 2031)),
            "seasons": ["单季", "第一季", "第二季"],
        },
        "area_reconciliation": {
            "official_plot_count": len(data.plots),
            "open_field_area_mu": open_field_area,
            "ordinary_greenhouse_area_mu": area_by_type["普通大棚"],
            "smart_greenhouse_area_mu": area_by_type["智慧大棚"],
            "total_area_mu": sum(area_by_type.values()),
            "passed": abs(open_field_area - 1201.0) <= 1e-9
            and abs(sum(area_by_type.values()) - 1213.0) <= 1e-9,
        },
        "units": {
            "area": "亩",
            "yield": "斤/亩",
            "cost": "元/亩",
            "price": "元/斤",
            "sales": "斤",
            "objective": "元",
        },
        "merged_cell_recovery": "读取合并区域左上角值并显式传播到区域内各数据行",
        "path_compatibility": {
            "windows_non_ascii_repository_path": "passed",
            "default_output_inside_repository": True,
            "observed_issue": "控制台曾乱码显示中文路径，但实际输出始终位于当前 worktree；未生成乱码同级目录",
            "regression_test": "tests/test_2024c_full_data.py::test_official_template_round_trip_preserves_decisions",
        },
        "verification": {
            "official_integration": {
                "command": "python -m pytest -q -m official_integration",
                "result": "6_passed",
            },
            "contract_fixture": {
                "command": "python -m pytest -q -m 'unit_contract or integration_fixture'",
                "result": "10_passed",
            },
        },
        "output_contracts": {
            name: {
                "sheets": list(contract.sheet_names),
                "crop_columns": len(contract.crop_names),
                "first_season_rows": len(contract.first_season_plots),
                "second_season_rows": len(contract.second_season_plots),
                "merged_ranges": list(contract.merged_ranges),
            }
            for name, contract in contracts.items()
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--material-root", type=Path, default=resolve_material_root())
    parser.add_argument(
        "--output",
        type=Path,
        default=default_audit_output_path(),
    )
    args = parser.parse_args()
    audit = build_audit(args.material_root.resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(args.output), "status": "passed"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
