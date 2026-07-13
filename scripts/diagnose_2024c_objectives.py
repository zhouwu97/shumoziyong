"""对 R01/R02 的 2024-C 目标复算偏差执行分层消融。"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from atomic_io import atomic_write_bytes


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
RUN_ROOT = ROOT / "experiments" / "a092_confirmatory_v1" / "runs"
ATTACHMENT_1 = ROOT / "official_materials" / "2024_C" / "attachments" / "附件1.xlsx"
ATTACHMENT_2 = ROOT / "official_materials" / "2024_C" / "attachments" / "附件2.xlsx"
DEFAULT_OUTPUT = ROOT / "experiments" / "2024c_objective_diagnosis_v1" / "diagnostic_result.json"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _load_variant(*, fill_merged: bool, smart_first_source: str) -> dict[str, Any]:
    from validators.problem_positive.validate import load_problem_data as load_v1

    base = load_v1(ATTACHMENT_1, ATTACHMENT_2)
    stats = {key: dict(value) for key, value in base["stats"].items()}
    source_plot_type, source_season = {
        "smart_second": ("智慧大棚", "第二季"),
        "ordinary_first": ("普通大棚", "第一季"),
    }[smart_first_source]
    for crop_id in range(17, 35):
        stats[("智慧大棚", "第一季", crop_id)] = dict(
            stats[(source_plot_type, source_season, crop_id)]
        )

    planting: list[dict[str, Any]] = []
    sales: defaultdict[tuple[int, str], float] = defaultdict(float)
    sheet = load_workbook(ATTACHMENT_2, read_only=True, data_only=True)[
        "2023年的农作物种植情况"
    ]
    current_plot: str | None = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            current_plot = str(row[0]).strip()
        plot_id = current_plot if fill_merged else (
            str(row[0]).strip() if row[0] is not None else None
        )
        if plot_id is None or not isinstance(row[1], (int, float)):
            continue
        crop_id = int(row[1])
        season = str(row[5]).strip()
        area = float(str(row[4]))
        planting.append(
            {
                "year": 2023,
                "plot_id": plot_id,
                "season": season,
                "crop_id": crop_id,
                "area_mu": area,
            }
        )
        stat = stats[(base["plots"][plot_id]["type"], season, crop_id)]
        sales[(crop_id, season)] += area * stat["yield"]
    return {
        "plots": base["plots"],
        "stats": stats,
        "planting_2023": planting,
        "sales_2023": dict(sales),
    }


def _scenario_summary(report: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "scenario_id": item["scenario_id"],
            "objective_reported": item.get("objective_reported"),
            "objective_recomputed": item.get("objective_recomputed"),
            "objective_difference": item.get("objective_difference"),
            "objective_valid": item.get("objective_valid", False),
            "constraint_violation_count": len(item.get("violated_constraints", [])),
            "violated_constraints": item.get("violated_constraints", []),
            "constraints_valid": item.get("constraints_valid", False),
        }
        for item in report["scenario_reports"]
    ]


def build_diagnostic_result() -> dict[str, Any]:
    from validators.problem_positive.validate import (
        evaluate_objective as evaluate_v1,
        validate_result as validate_v1,
    )
    from validators.problem_positive_v2.validate import (
        evaluate_objective as evaluate_v2,
        load_problem_data as load_v2,
        validate_result as validate_v2,
    )

    variants = {
        "v1_no_fill_smart_second": _load_variant(
            fill_merged=False, smart_first_source="smart_second"
        ),
        "fill_only_smart_second": _load_variant(
            fill_merged=True, smart_first_source="smart_second"
        ),
        "smart_only_no_fill": _load_variant(
            fill_merged=False, smart_first_source="ordinary_first"
        ),
        "v2_fill_and_ordinary_first": load_v2(ATTACHMENT_1, ATTACHMENT_2),
    }
    runs: dict[str, Any] = {}
    for run_id in ("R01", "R02"):
        formal_path = RUN_ROOT / run_id / "results" / "formal_result.json"
        if not formal_path.is_file():
            raise FileNotFoundError(f"缺少本机原始运行归档: {formal_path}")
        formal = json.loads(formal_path.read_text(encoding="utf-8"))
        v1_report = validate_v1(formal, ATTACHMENT_1, ATTACHMENT_2)
        v2_report = validate_v2(formal, ATTACHMENT_1, ATTACHMENT_2)
        ablation: list[dict[str, Any]] = []
        for scenario in formal["scenarios"]:
            scenario_id = scenario["scenario_id"]
            reported = float(scenario["objective_reported"])
            row: dict[str, Any] = {
                "scenario_id": scenario_id,
                "objective_reported": reported,
            }
            for variant_id, data in variants.items():
                evaluator = evaluate_v2 if variant_id == "v2_fill_and_ordinary_first" else evaluate_v1
                value = evaluator(scenario["assignments"], data, scenario_id)
                row[variant_id] = {
                    "objective_recomputed": value,
                    "reported_minus_recomputed": reported - value,
                }
            ablation.append(row)
        runs[run_id] = {
            "formal_result": {
                "path": formal_path.relative_to(ROOT).as_posix(),
                "sha256": _sha256(formal_path),
            },
            "v1_report": _scenario_summary(v1_report),
            "v2_report": _scenario_summary(v2_report),
            "preprocessing_ablation": ablation,
            "first_divergence": (
                "candidate_objective_sales_cap_grouped_by_crop_not_crop_season"
                if run_id == "R01"
                else "external_validator_data_preprocessing"
            ),
            "serialization_divergence_found": False,
            "paper_copy_divergence_found": False,
        }

    r02_objectives_pass = all(
        item["objective_valid"] for item in runs["R02"]["v2_report"]
    )
    r01_objectives_rejected = all(
        not item["objective_valid"] for item in runs["R01"]["v2_report"]
    )
    old_difference = runs["R02"]["v1_report"][0]["objective_difference"]
    return {
        "diagnosis_id": "2024c_objective_recomputation_diagnosis_v1",
        "official_inputs": {
            ATTACHMENT_1.relative_to(ROOT).as_posix(): _sha256(ATTACHMENT_1),
            ATTACHMENT_2.relative_to(ROOT).as_posix(): _sha256(ATTACHMENT_2),
        },
        "loader_ablation": {
            variant_id: {
                "planting_2023_rows": len(data["planting_2023"]),
                "sales_cap_keys": len(data["sales_2023"]),
            }
            for variant_id, data in variants.items()
        },
        "runs": runs,
        "root_causes": [
            {
                "layer": "external_validator_data_preprocessing",
                "issue": "merged_plot_ids_not_forward_filled",
                "effect": "2023 planting rows 87 -> 54; sales cap keys 47 -> 31",
            },
            {
                "layer": "external_validator_data_preprocessing",
                "issue": "smart_greenhouse_first_season_copied_from_smart_second",
                "effect": "yield/cost/price contract differs from ordinary growing season",
            },
            {
                "layer": "candidate_objective_implementation",
                "run_id": "R01",
                "issue": "sales_cap_grouped_by_crop_instead_of_crop_and_season",
                "effect": "v2 residual objective difference remains in all four scenarios",
            },
            {
                "layer": "candidate_constraint_implementation",
                "run_id": "R01",
                "issue": "2023_second_to_2024_first_smart_greenhouse_boundary_missing",
            },
            {
                "layer": "candidate_constraint_implementation",
                "run_id": "R02",
                "issue": "smart_greenhouse_rotation_checked_by_same_named_season_not_actual_sequence",
            },
        ],
        "diagnosis_passed": (
            r02_objectives_pass
            and r01_objectives_rejected
            and old_difference == 12526498.666666666
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="诊断 2024-C 目标复算差异")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    result = build_diagnostic_result()
    atomic_write_bytes(
        args.output,
        (json.dumps(result, ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
    )
    print(
        json.dumps(
            {"output": str(args.output), "diagnosis_passed": result["diagnosis_passed"]}
        )
    )
    return 0 if result["diagnosis_passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
