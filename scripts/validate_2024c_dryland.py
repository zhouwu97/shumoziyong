"""独立复算 2024-C 单季旱地基线；不调用候选求解器的目标函数。"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from jsonschema import Draft202012Validator
from openpyxl import load_workbook

from canonical_json import CANONICALIZATION_PRECISION, CANONICALIZATION_VERSION, canonical_sha256


ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = ROOT / "schemas" / "decision_variables.schema.json"
DRYLAND_TYPES = {"平旱地", "梯田", "山坡地"}
FEASIBILITY_TOLERANCE = 1e-9
OBJECTIVE_TOLERANCE = 1e-6
VALIDATION_PRECISION = "IEEE-754 float64"
DATA_EXTRACTION_CONTRACT = ROOT / "runtime_contracts" / "2024c_q1_dryland_extraction.json"


def _load_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError("决策变量根节点必须是对象")
    return value


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _validate_schema(value: dict[str, Any]) -> None:
    schema = _load_json(SCHEMA_PATH)
    errors = list(Draft202012Validator(schema).iter_errors(value))
    if errors:
        raise ValueError("决策变量格式无效：" + "；".join(error.message for error in errors))


def _price_midpoint(value: str) -> float:
    left, right = value.strip().split("-")
    return (float(left) + float(right)) / 2


def _source_tables(attachment_1: Path, attachment_2: Path) -> tuple[dict[str, float], dict[tuple[int, str], float]]:
    land_book = load_workbook(attachment_1, data_only=True, read_only=True)
    land = land_book["乡村的现有耕地"]
    capacities = {
        str(row[0]): float(row[2])
        for row in land.iter_rows(min_row=2, values_only=True)
        if row[1] in DRYLAND_TYPES and isinstance(row[2], (int, float))
    }
    stats_book = load_workbook(attachment_2, data_only=True, read_only=True)
    stats = stats_book["2023年统计的相关数据"]
    profits: dict[tuple[int, str], float] = {}
    for row in stats.iter_rows(min_row=2, values_only=True):
        crop_id, land_type, season = row[1], row[3], row[4]
        yield_per_mu, cost_per_mu, price = row[5], row[6], row[7]
        if (
            isinstance(crop_id, int)
            and 1 <= crop_id <= 15
            and land_type in DRYLAND_TYPES
            and season == "单季"
            and isinstance(yield_per_mu, (int, float))
            and isinstance(cost_per_mu, (int, float))
            and isinstance(price, str)
        ):
            profits[(crop_id, str(land_type))] = float(yield_per_mu) * _price_midpoint(price) - float(cost_per_mu)
    return capacities, profits


def validate_decision(
    decision_path: Path, attachment_1: Path, attachment_2: Path, material_manifest: Path
) -> dict[str, Any]:
    """从正式决策变量和官方附件独立重算利润、容量和定义域。"""
    decision = _load_json(decision_path)
    _validate_schema(decision)
    capacities, profits = _source_tables(attachment_1, attachment_2)
    used: dict[str, float] = {plot_id: 0.0 for plot_id in capacities}
    invalid = 0
    recomputed = 0.0
    details: list[str] = []
    for item in decision["assignments"]:
        plot_id, crop_id, area = item["plot_id"], item["crop_id"], float(item["area_mu"])
        key = (crop_id, _plot_type(attachment_1, plot_id)) if plot_id in capacities else None
        if key is None or key not in profits:
            invalid += 1
            details.append(f"非法地块-作物组合：{plot_id}-{crop_id}")
            continue
        used[plot_id] += area
        recomputed += area * profits[key]
    capacity_violation = max([0.0, *(area - capacities[plot_id] for plot_id, area in used.items())])
    # 此处使用未规范化的 float64 原始值进行数学判断；规范化只在全部通过后用于证据哈希。
    report = {
        "schema_version": "1.0.0",
        "validator": "optimization_validation_v1",
        "decision_variables_sha256": canonical_sha256(decision),
        "material_manifest_sha256": _sha256(material_manifest),
        "objective_reported": float(decision["objective_reported"]),
        "objective_recomputed": recomputed,
        "objective_abs_error": abs(float(decision["objective_reported"]) - recomputed),
        "max_capacity_violation": capacity_violation,
        "max_domain_violation": 0.0 if invalid == 0 else 1.0,
        "invalid_assignment_count": invalid,
        "feasible": invalid == 0 and capacity_violation <= FEASIBILITY_TOLERANCE and abs(float(decision["objective_reported"]) - recomputed) <= OBJECTIVE_TOLERANCE,
        "details": details,
        "validation_precision": VALIDATION_PRECISION,
        "feasibility_tolerance": FEASIBILITY_TOLERANCE,
        "objective_tolerance": OBJECTIVE_TOLERANCE,
        "canonicalization_version": CANONICALIZATION_VERSION,
        "canonicalization_precision": CANONICALIZATION_PRECISION,
        "data_extraction_contract_sha256": _sha256(DATA_EXTRACTION_CONTRACT),
    }
    return report


def _plot_type(attachment_1: Path, target: str) -> str:
    book = load_workbook(attachment_1, data_only=True, read_only=True)
    sheet = book["乡村的现有耕地"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == target:
            return str(row[1])
    raise ValueError(f"未知地块：{target}")
