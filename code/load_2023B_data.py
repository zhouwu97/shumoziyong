from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

NM_TO_M = 1852.0
DEFAULT_ATTACHMENT = Path("official_materials/2023_B/attachments/附件.xlsx")


@dataclass(frozen=True)
class DepthGrid:
    x_nm: np.ndarray
    y_nm: np.ndarray
    depth_m: np.ndarray

    @property
    def x_m(self) -> np.ndarray:
        return self.x_nm * NM_TO_M

    @property
    def y_m(self) -> np.ndarray:
        return self.y_nm * NM_TO_M

    def summary(self) -> dict:
        return {
            "rows": int(self.depth_m.shape[0]),
            "cols": int(self.depth_m.shape[1]),
            "x_range_nm": (float(np.min(self.x_nm)), float(np.max(self.x_nm))),
            "y_range_nm": (float(np.min(self.y_nm)), float(np.max(self.y_nm))),
            "x_range_m": (float(np.min(self.x_m)), float(np.max(self.x_m))),
            "y_range_m": (float(np.min(self.y_m)), float(np.max(self.y_m))),
            "depth_range_m": (float(np.min(self.depth_m)), float(np.max(self.depth_m))),
            "depth_mean_m": float(np.mean(self.depth_m)),
            "has_nan": bool(np.isnan(self.depth_m).any()),
        }


def load_depth_grid(path: Path | str = DEFAULT_ATTACHMENT) -> DepthGrid:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb.active

    x_values = [cell.value for cell in ws[2][2:] if cell.value is not None]
    y_values = []
    depth_rows = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        y = row[1]
        values = row[2 : 2 + len(x_values)]
        if y is None or all(value is None for value in values):
            continue
        y_values.append(y)
        depth_rows.append(values)

    x_nm = np.array(x_values, dtype=float)
    y_nm = np.array(y_values, dtype=float)
    depth_m = np.array(depth_rows, dtype=float)

    if depth_m.shape != (len(y_nm), len(x_nm)):
        raise ValueError("水深网格尺寸与坐标轴长度不一致")
    if np.isnan(depth_m).any():
        raise ValueError("水深网格包含空值，需先清洗后再建模")
    if np.any(depth_m <= 0):
        raise ValueError("水深必须为正值")

    return DepthGrid(x_nm=x_nm, y_nm=y_nm, depth_m=depth_m)


def main() -> None:
    grid = load_depth_grid()
    for key, value in grid.summary().items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
