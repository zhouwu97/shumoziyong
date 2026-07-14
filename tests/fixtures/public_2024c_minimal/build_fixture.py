"""生成公开、最小的 2024-C Excel 回归样本。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent


def write_attachment_1() -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "乡村的现有耕地"
    sheet.append(["地块名称", "地块类型", "地块面积/亩"])
    sheet.append(["A1", "平旱地", 10])
    sheet.append(["F1", "智慧大棚", 0.6])
    book.save(ROOT / "附件1.xlsx")


def write_attachment_2() -> None:
    book = Workbook()
    stats = book.active
    stats.title = "2023年统计的相关数据"
    stats.append(["序号", "作物编号", "作物名称", "地块类型", "种植季次", "亩产量/斤", "种植成本/(元/亩)", "销售单价/(元/斤)"])
    stats.append([1, 1, "黄豆", "平旱地", "单季", "100", "10.5", "2.00-4.00"])
    for crop_id in range(17, 35):
        stats.append(
            [crop_id, crop_id, f"蔬菜{crop_id}", "普通大棚", "第一季", 50, 20, "7-9"]
        )
    stats.append([100, 17, "番茄", "智慧大棚", "第二季", 40, 30, "19-21"])

    planting = book.create_sheet("2023年的农作物种植情况")
    planting.append(["地块名称", "作物编号", "作物名称", "作物类型", "种植面积/亩", "种植季次"])
    planting.append(["A1", 1, "黄豆", "粮食", 4, "单季"])
    planting.append([None, 1, "黄豆", "粮食", 1, "单季"])
    planting.append(["F1", 17, "番茄", "蔬菜", 0.6, "第二季"])
    planting.append([None, None, None, None, None, None])
    planting.merge_cells("A2:A3")
    book.save(ROOT / "附件2.xlsx")


if __name__ == "__main__":
    write_attachment_1()
    write_attachment_2()
